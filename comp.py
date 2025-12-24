#!/usr/bin/env python3
# This script fetches specific financial labels from the SEC data, now supporting
# single-tag fallbacks, two-component, three-component, and four-component
# calculated fallbacks.
# The script now outputs two consolidated Excel files:
# 1. <Name>_Full_Data.xlsx (Raw and Calculated sheets for ALL tickers)
# 2. <Name>_Comparison_Data.xlsx (Peer comparison, CAGR, and price summary sheets)

import requests
import pandas as pd
import os
import time
import yfinance as yf
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from datetime import date
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import pytz
import numpy as np
import json
import copy

# =========================================================================
# === IMPORTANT: REQUIRED LIBRARIES ===
# Run the following command to install them in your Mac Terminal:
#
# pip install lxml openpyxl yfinance pandas requests numpy
#
# To run this script after saving it as 'scraper.py':
# 1. cd /path/to/where/you/saved/it
# 2. python3 scraper.py
# 3. The script will now process all 31 hardcoded tickers and create the two consolidated Excel books.
#---------
#NOTES:
# Fair Value Screen..... Comparitive Data Screen for the fair value screen data will look off due to the closing price being not scraped properly well. Fix this later



# =========================================================================

# --- SEC API Configuration ---
# The SEC requires a User-Agent header for all API requests.
HEADERS = {
    'User-Agent': 'StockAnalysisResearch/1.0 (contact: github.com/caseyskan-rgb)'
    
}
RISK_FREE_RATE = 0.04          # 4% risk-free rate (e.g., 10-year U.S. Treasury)
EQUITY_RISK_PREMIUM = 0.055    # 5.5% equity market risk premium
MIN_COST_OF_DEBT = 0.02        # 2% minimum cost of debt

POLYGON_API_KEY = os.getenv("POLYGON_API_KEY")
if not POLYGON_API_KEY:
    raise RuntimeError("Polygon API key not found in environment variables")

# --- Peer Comparison and CAGR Labels ---
# These are the 9 metrics defined by the user for the combined CAGR sheet.
CAGR_LABELS_FOR_EXPORT = [
    "Total Revenue",
    "Net Income",
    "Total Assets",
    "Total Common Shares Outstanding",
    "Cash From Operations",
    "Earnings Per Share (EPS)",
    "Gross Profit Margin (%)",
    "Net Profit Margin (%)",
    "ROIC (Heavy)"
]

# --- Hardcoded SEC Labels and Excel Names (RELEVANT_LABELS - UNCHANGED) ---
RELEVANT_LABELS = {
    "Total Revenue": [
        "Revenues",
        "RevenueFromContractWithCustomerExcludingAssessedTax",
        "SalesRevenueNet"
    ],
    "Net Income": [
        "NetIncomeLoss",
        "IncomeLossFromContinuingOperations"
    ],
    "Total Assets": [
        "Assets",
        "AssetsTotal"
    ],
    "Total Liabilities": [
        "Liabilities",
        "LiabilitiesTotal",
        ["CALCULATION:", "LiabilitiesAndStockholdersEquity", "StockholdersEquity", "-"],
        ["CALCULATION:", "LiabilitiesCurrent", "+", "LiabilitiesNoncurrent"]
    ],
    "Preferred Stock": [
        "PreferredStockValueOutstanding",
        "PreferredStockValue"
    ],
    "Preferred Stock Issued": [
        "PreferredStockSharesIssued",
    ],
    "Total Common Shares Outstanding": [
        "CommonStockSharesOutstanding",
        "WeightedAverageNumberOfSharesOutstandingBasic"
    ],
    "Short Term Debt (STD)": [
        "LongTermDebtCurrent",
        "ConvertibleSeniorNotesCurrent",
        "NotesPayableCurrent",
        ["CALCULATION:", "LongTermDebt", "LongTermDebtNoncurrent", "-"],
        ["CALCULATION:", "LongTermDebtCurrent", "+", "ConvertibleSeniorNotesCurrent", "+", "NotesPayableCurrent"],
        ["CALCULATION:", "ConvertibleSeniorNotesCurrent", "+", "NotesPayableCurrent", "+", "OperatingLeaseLiabilityCurrent"]
    ],
    "Cash & Cash Equivelance": [
        "CashAndCashEquivalentsAtCarryingValue",
        "CashAndCashEquivalents",
        "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"
    ],
    "Minority Interest": [
        "MinorityInterest",
        "NoncontrollingInterest",
        "NetIncomeLossAttributableToNoncontrollingInterest",
        "NetIncomeLossAttributableToNoncontrollingInterestIncludingPortionAttributableToRedeemableNoncontrollingInterest",
        "NoncontrollingInterestInVariableInterestEntity"
    ],
    "Long Term Debt (LTD)": [
        "ConvertibleSeniorNotes",
        "LongTermDebt",
        ["CALCULATION:", "ConvertibleSeniorNotes", "+", "LongTermDebt", "+", "OperatingLeaseLiability"],
        ["CALCULATION:", "ConvertibleSeniorNotes", "OperatingLeaseLiability", "+"]
    ],
    "Operating Lease Liability": [
        "OperatingLeaseLiability"
    ],
    # --- START LEASE FOOTNOTE ADDITIONS (For WADR and ROU Asset) ---
    "Operating Lease ROU Asset": [
        "OperatingLeaseRightOfUseAsset",
        "OperatingLeaseRightOfUseAsset"
    ],
    "Weighted Average Lease Discount Rate": [
        "WeightedAverageDiscountRate",
    ],
    # --- END LEASE FOOTNOTE ADDITIONS ---
    "Operating Income": [
        "OperatingIncomeLoss"
    ],
    "Depreciation, Depletion, and Amortization": [
        "DepreciationDepletionAndAmortization",
        # Original 2-component structure preserved (CALCULATION: 4 elements)
        ["CALCULATION:", "DepreciationAndAmortization", "Depletion", "+"],
        ["CALCULATION:", "DepreciationAndAmortization", "DepletionExpense", "+"],
        # Example 3-Component structure (CALCULATION: 6 elements)
        ["CALCULATION:", "Depreciation", "+", "AmortizationOfIntangibleAssets", "+", "Depletion"],
        "DepreciationAndAmortization",
        ["CALCULATION:", "AmortizationOfIntangibleAssets", "Depreciation", "+"],
        "Depreciation"
    ],
    "Cost of Goods Sold": [
        "CostOfGoodsAndServicesSold",
        "CostOfGoodsSold",
        "CostOfRevenue"
    ],
    "Gross Profit": [
        "GrossProfit",
        ["CALCULATION:", "Revenues", "CostOfGoodsAndServicesSold", "-"],
        ["CALCULATION:", "Revenues", "CostOfRevenue", "-"],
        ["CALCULATION:", "Revenues", "CostOfGoodsSold", "-"],
        ["CALCULATION:", "RevenueFromContractWithCustomerExcludingAssessedTax", "CostOfRevenue", "-"],
        ["CALCULATION:", "RevenueFromContractWithCustomerExcludingAssessedTax", "CostOfGoodsAndServicesSold", "-"],
        ["CALCULATION:", "RevenueFromContractWithCustomerExcludingAssessedTax", "SalesRevenueNet", "-"]
    ],
    "Operating Expense (EBIT)": [
        "OperatingExpense",
        ["CALCULATION:", "GrossProfit", "OperatingIncomeLoss", "-"],
        # 3-Component Calculations (Revenues - COGS - OperatingIncome = Operating Expense)
        ["CALCULATION:", "Revenues", "-", "CostOfGoodsAndServicesSold", "-", "OperatingIncomeLoss"],
        ["CALCULATION:", "Revenues", "-", "CostOfRevenue", "-", "OperatingIncomeLoss"],
        ["CALCULATION:", "Revenues", "-", "CostOfGoodsSold", "-", "OperatingIncomeLoss"],
        ["CALCULATION:", "RevenueFromContractWithCustomerExcludingAssessedTax", "-", "CostOfRevenue", "-", "OperatingIncomeLoss"],
        ["CALCULATION:", "RevenueFromContractWithCustomerExcludingAssessedTax", "-", "CostOfGoodsAndServicesSold", "-", "OperatingIncomeLoss"],
        ["CALCULATION:", "RevenueFromContractWithCustomerExcludingAssessedTax", "-", "CostOfGoodsSold", "-", "OperatingIncomeLoss"],
        ["CALCULATION:", "ResearchAndDevelopmentExpense", "+", "GeneralAndAdministrativeExpense", "+", "SellingAndMarketingExpense"],
        ["CALCULATION:", "ResearchAndDevelopmentExpense", "+", "SellingGeneralAndAdministrativeExpense", "+", "OtherOperatingExpenses"],
        ["CALCULATION:", "ResearchAndDevelopmentExpense", "+", "GeneralAndAdministrativeExpense", "+", "SellingExpense"]
    ],
    "Shareholder/Stockholder Equity": [
        "StockholdersEquity",
        "ShareholdersEquity",
        "Equity"
    ],
    "Net PP&E": [
        # Syntax Fixed: Cleaned up list
        "PropertyPlantAndEquipmentNet",
        "PropertyPlantAndEquipmentAndFinanceLeaseRightOfUseAssetAfterAccumulatedDepreciationAndAmortization",
        ["CALCULATION:", "PropertyPlantAndEquipment", "AccumulatedDepreciationDepletionAndAmortization", "-"]
    ],
    "Interest Expense": [
        "InterestExpense",
        "InterestExpenseDebt",
        "InterestPaidNet",
        "InterestExpenseOperating",
    ],
    "Income Tax Expense": [
        "IncomeTaxExpenseBenefit"
    ],
    "Income Before Tax": [
        "IncomeLossBeforeIncomeTaxExpenseBenefit",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxExpenseBenefit"
    ],
    
    ### END NEW VALUATION INGREDIENTS ###
    
    "Inventory": [
        "InventoryNet",
        "Inventory"
    ],
    "Accounts Receivable": [
        ["CALCULATION:", "AccountsReceivableNetCurrent", "+", "AccountsReceivableNetNoncurrent"],
        "AccountsReceivableNetCurrent",
        "AccountsReceivableNetNoncurrent"
    ],
    "Short Term Investments": [
        "MarketableSecuritiesCurrent",
        "ShortTermInvestments",
        ["CALCULATION:", "CashCashEquivalentsAndShortTermInvestments", "CashAndCashEquivalentsAtCarryingValue", "-"],
        ["CALCULATION:", "MarketableSecurities", "MarketableSecuritiesNoncurrent", "-"]
    ],
    "Current Assets": [
        "AssetsCurrent",
        ["CALCULATION:", "Assets", "AssetsNoncurrent", "-"],
        ["CALCULATION:", "Assets", "NoncurrentAssets", "-"]
    ],
    "Current Liabilities": [
        "LiabilitiesCurrent",
        ["CALCULATION:", "Liabilities", "LiabilitiesNoncurrent", "-"]
    ],
    "Goodwill": [
        "Goodwill"
    ],
    "Net Intangible Assets": [
        "IntangibleAssetsNetExcludingGoodwill",
        "IntangibleAssetsOtherThanGoodwillNet",
        "FiniteLivedIntangibleAssetsNet", # Has an estimated useful life
        # Corrected possible previous typo for "CALCULATION:"
        ["CALCULATION:", "IntangibleAssets", "Goodwill", "-"],
    ],
    "Notes Payable": [
        "LongTermNotesPayable",
        "NotesPayable",
        "NotesPayableCurrent"
    ],
    "Commercial Papers": [
        "CommercialPaper"
    ],
    "Current Bank Overdraft": [
        "CashOverdraft",
        "CurrentBankOverdraft",
        "BankOverdraft"
    ],
    "Convertible Senior Notes": [
        "ConvertibleSeniorNotes",
        ["CALCULATION:", "ConvertibleSeniorNotesNoncurrent", "+", "ConvertibleSeniorNotesCurrent"],
        "ConvertibleSeniorNotesNoncurrent",
        "ConvertibleSeniorNotesCurrent"
    ],
    
    "Selling & Marketing Expense": [
        "SellingAndMarketingExpense"
    ],
    "Research & Development Expense": [
        "ResearchAndDevelopmentExpense"
    ],
    "General & Administrative Expense": [
        "GeneralAndAdministrativeExpense"
    ],
    "Marketing Expense": [
        "MarketingExpense"
    ],
    "Other Operating Expense": [
        "OtherOperatingExpense"
    ],
    
    #----------------------------------------------------------------------------------
    # THis is where the second portion of the stock analysis begins...the source(Debt,Equity,Operations) and use of fund (CFO, CFF, CFI)
    "Cash From Operations": [
        "NetCashProvidedByUsedInOperatingActivities", # if this one is used and the value is negative, then it is a cash outflow toward operations.
        "CashAndCashEquivalentsProvidedByUsedInOperatingActivities",
        "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations",
        "NetCashFlowsFromUsedInOperatingActivities" # Source of Fund (1)
    ],
    "Proceed From Issuance of New Stock / Equity": [ #Source of Fund (2)
        "ProceedsFromIssuanceOrSaleOfEquity",
        "ProceedsFromIssuanceOfCommonStock",
        "ProceedsFromIssuanceOfPreferredStock",
        "ProceedsFromStockOptionsExercised"
    ],
    "Proceeds From Issuance Of Debt": [
        "ProceedsFromDebtNetOfIssuanceCosts",
        "ProceedsFromIssuanceOfDebt"
    ],
    "Proceed From Long Term Debt": [ #Source of Fund (3)
        "ProceedsFromIssuanceOfLongTermDebt",
        "ProceedsFromLongTermDebt",
        "ProceedsFromBorrowings",
        "ProceedsFromDebtMaturingInMoreThanThreeMonths"
    ],
    "Proceed From Short Term Debt": [
        "ProceedsFromShortTermDebt"
    ],
    
    # Here is cash outflow
    
    "Payments For Repurchase of Stock": [ #Financing Outflow (1)
        "PaymentsForRepurchaseOfCommonStock",
        "PaymentsForRepurchaseOfEquity"
    ],
    "Dividend Common Stock Cash": [ #Financing Outflow (2)
        "DividendCommonStockCash",
        "PaymentsOfDividendsCommonStock",
        "PaymentsOfDividends",
    ],
    
    "Repayment Of Debt": [ #Financing Outflow (3)
        "RepaymentsOfDebtAndCapitalLeaseObligations",
        "PaymentsForRepaymentOfDebt"
    ],
    "Repayment Of Long Term Debt": [ #Addition for #3 with STD Incase not there
        "RepaymentsOfLongTermDebt",
        "RepaymentsOfDebtMaturingInMoreThanThreeMonths",
    ],
    "Repayment Of Short Term Debt": [ #Addition for #3 with LTD Incase not there
        "RepaymentsOfShortTermDebt"
    ],
    "Finance Lease Principal Payments": [
        "FinanceLeasePrincipalPayments" #Normally this is included in Repayment of Debt, so this is just incase no label is found for the other one
        
        #--Total Outflow Investment = CapEx(Inttang and Tang)+Payments To Aquire Investment + M$A
    ],
    "SegmentExpenditureAdditionToLongLivedAssets": [
        "SegmentExpenditureAdditionToLongLivedAssets" #OVERALL CapEx (Physical and Intangible)
    ],
    "Payments To Acquire Investments": [ # Use of Fund Investing (1)
        "PaymentsToAcquireInvestments",
        "PaymentsToAcquireMarketableSecurities"
    ],
    "Pay. Acquire Other Investment": [
        "PaymentsToAcquireOtherInvestments"
    ],
    
    "Payments To Acquire PP&E": [ #Investing Use Add (2) (Physical CapEx)
        "PaymentsToAcquirePropertyPlantAndEquipment",
        "PaymentsToAcquireProductiveAssets",
    ],
    "Payment To Acquire Productive Assets": [ #Backup for PaymentToAquirePP&E
        "PaymentsToAcquireProductiveAssets" #Investing Use Add (2)
    ],
    "Payments For Software": [ #Investment Use Add (BACKUP...Just to see)
        "PaymentsForSoftware"
    ],
    "Payments To Acquire Intangible Assets": [ #Investment Use Add (3) (Intangible CapEx)
        "PaymentsToAcquireIntangibleAssets"
        "PaymentForSoftware"
    ],
    "Payments To Acquire Businesses Net Of Cash Acquired": [
        "PaymentsToAcquireBusinessesNetOfCashAcquired" #Investing Use Add (M&A Activity) (4)
    ],
    # ADDED Interest Expense Non-Operating back as a raw tag to ensure it's scraped first
    "Interest Expense Non-Operating": [
        "InterestExpenseNonoperating"
    ],
    "Prepaid Expenses": [
        # Note: This is an exception. It's defined here for scraping fallbacks,
        # but the actual calculation runs below in CALCULATED_EQUATIONS.
        # This prevents the scraper from completely missing it if the raw components
        # are not found elsewhere, ensuring it is zero-filled properly.
        "PrepaidExpenseAndOtherAssetsCurrent",
        "OtherAssetsCurrent"
    ],
} # The multi-variable calculations were CUT from here and moved below.

# --- Python-Calculated Static Assumptions (STATIC_ASSUMPTIONS - UNCHANGED) ---
STATIC_ASSUMPTIONS = {
    "Min Operating Cash (5-Year Lookback)": { #Target Percentage
        "components": ["Cash & Cash Equivelance", "Total Revenue"],
        "operation": "MIN_RATIO_N_YEARS", # New operation to find the historical minimum ratio
        "years": 5,           # Lookback period
        "description": "Minimum Cash/Revenue Ratio over the last N years, used as the Non-Operating Cash assumption for Invested Capital."
        #If needed just add "multiplier": 100, to get a percentage
        }
}

# --- Python-Calculated Metrics (CALCULATED_EQUATIONS - UNCHANGED) ---
CALCULATED_EQUATIONS = {
    
    # =======================================================
    # === LEVEL 1: BASE CALCULATIONS (MUST RUN FIRST) ===
    # =======================================================
    "Interest Expense Consolidated": {
        "components": ["Interest Expense", "Interest Expense Non-Operating"],
        "operation": "CONSOLIDATE_FILL",
        "description": "Fills missing values in Interest Expense with the values of Non-Operating Interest Expense"
    },
    "Free Cash Flow": {
        "components": ["Cash From Operations","Payments To Acquire PP&E"],
        "operation": "-",
        "description": "Cash From Operation - Payment To Acquire PP&E"
    },
    "Average Total Assets": {
        "components": ["Total Assets"],
        "operation": "AVERAGE_PRIOR",
        "description": "(Current Year Assets + Prior Year Assets) / 2"
    },
    "Average Shareholder Equity": {
        "components": ["Shareholder/Stockholder Equity"],
        "operation": "AVERAGE_PRIOR",
        "description": "(Current Year Equity + Prior Year Equity) / 2"
    },
    
    # =======================================================
    # === LEVEL 2: MOVED PRE-CALCULATIONS (Building Blocks) ===
    # =======================================================
    "Book Value": {
        "components": ["Total Assets", "Total Liabilities"],
        "operation": "-",
        "description": "Total Assets - Total Liabilities"
    },
    "Tangible Book Value": {
        # Note: This is now a three-component formula since 'Book Value' is a component.
        "components": ["Book Value", "Preferred Stock", "Goodwill", "Net Intangible Assets"],
        "operation": "-",
        "description": "Book Value - Preferred Stock - Goodwill - Net Intangible Assets"
    },
    "Interest Bearing Debt": {
        "components": ["Short Term Debt (STD)", "Long Term Debt (LTD)", "Operating Lease Liability"],
        "operation": "+",
        "description": "Short Term Debt (STD) + Long Term Debt (LTD) + Operating Lease Liability"
    },
    "Net Debt": {
        "components": ["Interest Bearing Debt", "Cash & Cash Equivelance"],
        "operation": "-",
        "description": "Interest Bearing Debt - Cash & Cash Equivelance"
    },
    "Market Cap": {
        "components": ["Closing Price (USD)", "Total Common Shares Outstanding"],
        "operation": "*",
        "description": "Closing Price (USD) * Total Common Shares Outstanding"
    },
    "Enterprise Value": {
        # Note: This is now a multi-component formula that uses other calculated metrics
        "components": ["Market Cap", "Net Debt", "Preferred Stock", "Minority Interest"],
        "operation": "+", # Assuming all are added to Market Cap
        "description": "Market Cap + Net Debt + Preferred Stock + Minority Interest"
    },
    "EBITDA": {
        "components": ["Operating Income", "Depreciation, Depletion, and Amortization"],
        "operation": "+",
        "description": "Operating Income + Depreciation, Depletion, and Amortization"
    },
    "EBIT": {
        "components": ["Gross Profit", "Operating Expense (EBIT)"],
        "operation": "-",
        "description": "Gross Profit - Operating Expense (EBIT)"
    },
    "Prepaid Expenses": {
        # This calculation is run as a fallback/check after the initial scraping,
        # using the raw tags that were scraped and zero-filled.
        "components": ["PrepaidExpenseAndOtherAssetsCurrent", "OtherAssetsCurrent"],
        "operation": "-",
        "description": "PrepaidExpenseAndOtherAssetsCurrent - OtherAssetsCurrent"
    },
    # Intermediate Light ROIC components:
    "Accounts Recievable + Inventory + Prepaid Expenses" : {
        "components": ["Accounts Receivable", "Inventory", "Prepaid Expenses"],
        "operation": "+",
        "description": "Accounts Receivable + Inventory + Prepaid Expenses"
    },
    "Notes Payable + Current Bank Overdraft + Commercial Papers": {
        "components": ["Notes Payable", "Commercial Papers", "Current Bank Overdraft"],
        "operation": "+",
        "description": "Notes Payable + Commercial Papers + Current Bank Overdraft"
    },
    "Net Intangible Assets + Net PP&E + Goodwill": {
        "components": ["Net Intangible Assets", "Net PP&E", "Goodwill"],
        "operation": "+",
        "description": "Net Intangible Assets + Net PP&E + Goodwill"
    },
    "Net Intangible Assets + Net PP&E + Goodwill + Operating Lease ROU Asset": {
        "components": ["Net Intangible Assets + Net PP&E + Goodwill", "Operating Lease ROU Asset"],
        "operation": "+",
        "description": "Net Intangible Assets + Net PP&E + Goodwill + Operating Lease ROU Asset"
    },
    "Cash and Cash Equivalent + Short Term Investments + Accounts Recievable": {
        "components": ["Cash & Cash Equivelance", "Short Term Investments", "Accounts Receivable"],
        "operation": "+",
        "description": "Cash & Cash Equivelance + Short Term Investments + Accounts Recievable"
    },

    # =======================================================
    # === LEVEL 3: CORE RATIOS (Depends on Levels 1 & 2) ===
    # =======================================================
    "Net Working Capital (Heavy)": {
        "components": ["Current Assets", "Current Liabilities"],
        "operation": "-",
        "description": "Current Assets - Current Liabilities"
    },
    "Invested Capital (Heavy)": {
        "components": ["Net PP&E","Net Working Capital (Heavy)"],
        "operation": "+",
        "description": "Net PP&E + Net Working Capital (Heavy)"
    },
    "Asset Turnover Ratio": {
        "components": ["Total Revenue", "Average Total Assets"],
        "operation": "/",
        "description": "Total Revenue / Average Total Assets"
    },
    "Operating Margin": {
        "components": ["EBIT", "Total Revenue"],
        "operation": "/",
        "description": "(EBIT /Total Revenue)*100"
    },
    "Financial Leverage": {
        "components": ["Average Total Assets","Average Shareholder Equity"],
        "operation": "/",
        "description": "Average Total Assets / Average Shareholder Equity"
    },
    "EBT": {
        "components": ["EBIT","Interest Expense Consolidated"],
        "operation": "-",
        "description": "EBIT - Interest Expenses Consolidated"
    },
    "Tax Rate": {
        "components": ["Income Tax Expense","EBT"],
        "operation": "/",
        "description": "Income Tax Expense / EBT"
    },
    "Tax Retention Ratio": {
        "components": [1,"Tax Rate"],
        "operation": "-",
        "description": " 1 - Tax Rate"
    },
    # --- START NOPAT ADJUSTMENT (to use WADR) ---
    "Embedded Lease Interest": {
        # NOPAT Adjustment: Operating Lease Liability * WADR
        "components": ["Operating Lease Liability", "Weighted Average Lease Discount Rate"],
        "operation": "*",
        "description": "Operating Lease Liability * Weighted Average Lease Discount Rate"
    },
    "Net Operating Profit After Taxes (Heavy)": {
        "components": ["Operating Income", "Tax Retention Ratio",],
        "operation": "*",
        "description": "(Operating Income * Tax Retention Ratio)"
        # Note: The * operation is handled first, then the result of that is added to the 3rd component
    },
    "Net Operating Profit After Taxes (Light)": {
        "components": ["Net Operating Profit After Taxes (Heavy)","Embedded Lease Interest"],
        "operation": "+",
        "description": "(Operating Income * Tax Retention Ratio) + Embedded Lease Interest"
    },
    # --- END NOPAT ADJUSTMENT ---
    "Tax Burden": {
        "components": ["Net Income", "EBT"],
        "operation": "/",
        "description": "Net Income / EBT"
    },
    "Interest Burden": {
        "components": ["EBT","EBIT"],
        "operation": "/",
        "description": "EBT / EBIT"
    },
    
    # --- PROFITABILITY MARGINS & MULTIPLES (Level 3.5) ---
    "Gross Profit Margin (%)": {
        "components": ["Gross Profit", "Total Revenue"],
        "operation": "/",
        "multiplier": 100,
        "description": "(Gross Profit / Total Revenue) * 100"
    },
    "Net Profit Margin (%)": {
        "components": ["Net Income", "Total Revenue"],
        "operation": "/",
        "multiplier": 100,
        "description": "(Net Income / Total Revenue)*100"
    },
    "Operating Profit Margin (%)": {
        "components": ["EBIT", "Total Revenue"],
        "operation": "/",
        "multiplier": 100,
        "description": "(EBIT /Total Revenue)*100"
    },
    
    # --- ROIC (Heavy) ---
    "ROIC (Heavy)": {
        "components": ["Net Operating Profit After Taxes (Heavy)","Invested Capital (Heavy)"],
        "operation": "/",
        "description": "NOPAT / Invested Capital Heavy"
    },
    
    # --- ROIC (Light) CHAIN (Depends on Static Assumption 'Min Operating Cash...') ---
    "Operating Cash": {
        # This requires 'Min Operating Cash (5-Year Lookback)' to be pre-calculated in a separate step
        "components": ["Min Operating Cash (5-Year Lookback)","Total Revenue"],
        "operation": "*",
        "description": "Target Percent * Total Revenue"
    },
    "Excess Cash": {
        "components": ["Cash & Cash Equivelance","Operating Cash"],
        "operation": "-",
        "description": "Cash and Cash Equivelance + Operating Cash"
    },
    "Excess Cash + Short Term Investments": {
        "components": ["Excess Cash","Short Term Investments"],
        "operation": "+",
        "description": "Excess Cash + Short Term Investments (NOCA)"
    },
    
    # --- START MODIFICATION FOR OPERATING CURRENT ASSETS (OCA) ---
    # The metric "Current Assets (light)" is removed as it is structurally incorrect.
    # The new OCA calculation is simplified to: Total Current Assets - NOCA.
    "Operating Current Assets": {
        "components": ["Current Assets","Excess Cash + Short Term Investments"],
        "operation": "-",
        "description": "Current Assets - (Excess Cash + Short Term Investments) -> Total Current Assets - NOCA"
    },
    # --- END MODIFICATION FOR OPERATING CURRENT ASSETS (OCA) ---
    
    "Non Interest Bearing Current Liabilitites": {
        "components": ["Current Liabilities","Notes Payable + Current Bank Overdraft + Commercial Papers"],
        "operation": "-",
        "description": "Current Liabilities - Interest Bearing Current Liabilities (IBCL)"
    },
    "Net Working Capital (Light)": {
        "components": ["Operating Current Assets","Non Interest Bearing Current Liabilitites"],
        "operation": "-",
        "description": "Operating Current Assets - NIBCL"
    },
        
    "Invested Capital (Light)": {
        "components": ["Net Intangible Assets + Net PP&E + Goodwill + Operating Lease ROU Asset","Net Working Capital (Light)"],
        "operation": "+",
        "description": "Net Intangible Assets + Net PP&E + Goodwill + Operating Lease ROU Asset + Net Working Capital (Light)"
    },
    "ROIC (Light)": {
        "components": ["Net Operating Profit After Taxes (Light)","Invested Capital (Light)"],
        "operation": "/",
        "description": "Net Operating Profit After Taxes (Light) / Invested Capital (Light)"
    },
    
    # --- ROA CHAIN (Level 3) ---
    "Interest Expense Consolidated * Tax Burden": {
        "components": ["Interest Expense Consolidated","Tax Burden"],
        "operation": "*",
        "description": "Interest Expense Consolidated * Tax Burden"
    },
    "Net Income + Interest Expense Consolidated * Tax Burden": {
        "components": ["Interest Expense Consolidated * Tax Burden","Net Income"],
        "operation": "+",
        "description": "(Interest Expense Consolidated * Tax Burden) + Net Income"
    },
    "Return On Assets": {
        "components": ["Net Income + Interest Expense Consolidated * Tax Burden","Average Total Assets"],
        "operation": "/",
        "description": "(Net Income + (Interest Expense Consolidated * Tax Burden))/ Average Total Assets"
    },
    
    # --- ROE (DuPont) CHAIN (Level 3) ---
    "Tax Burden * Interest Burden": {
        "components": ["Tax Burden","Interest Burden"],
        "operation": "*",
        "description": "Tax Burden * Interest Burden"
    },
    "Tax Burden * Interest Burden * Operating Margin" : {
        "components": ["Tax Burden * Interest Burden","Operating Margin"],
        "operation": "*",
        "description": "Tax Burden * Interest Burden * Operating Margin"
    },
    "Tax Burden * Interest Burden * Operating Margin * Asset Turnover Ratio": {
        "components": ["Tax Burden * Interest Burden * Operating Margin","Asset Turnover Ratio"],
        "operation": "*",
        "description": "Tax Burden * Interest Burden * Operating Margin * Asset Turnover Ratio * Asset Turnover Ratio"
    },
    "Return On Equity": {
        "components": ["Tax Burden * Interest Burden * Operating Margin * Asset Turnover Ratio","Financial Leverage"],
        "operation": "*",
        "description": "Tax Burden * Interest Burden * Operating Margin * Asset Turnover Ratio * Financial Leverage"
    },

    # --- Valuation Multiples (Level 3) ---
    "Tangible Book Value Per Share": {
        "components": ["Tangible Book Value", "Total Common Shares Outstanding"],
        "operation": "/",
        "description": "Tangible Book Value / Total Common Shares Outstanding"
    },
    "Price To Book (P/B Ratio)": {
        "components": ["Closing Price (USD)", "Tangible Book Value Per Share"],
        "operation": "/",
        "description": "Closing Price (USD) / Tangible Book Value Per Share"
    },
    "EV / EBITDA": {
        "components": ["Enterprise Value", "EBITDA"],
        "operation": "/",
        "description": "Enterprise Value / EBITDA"
    },
    "EV / Share": {
        "components": ["Enterprise Value", "Total Common Shares Outstanding"],
        "operation": "/",
        "description": "Enterprise Value / Total Common Shares Outstanding"
    },
    # --- P/E Calculation Components ---
    "Earnings Per Share (EPS)": {
        "components": ["Net Income", "Total Common Shares Outstanding"],
        "operation": "/",
        "description": "Net Income / Total Common Shares Outstanding"
    },
    "P/E Ratio (Calculated)": {
        "components": ["Closing Price (USD)", "Earnings Per Share (EPS)"],
        "operation": "/",
        "description": "Closing Price (USD) / Earnings Per Share (EPS)"
    },
    
    # --- CAGR and PEG Ratio (Now all 9 are defined here for robustness) ---
    "Total Revenue CAGR (5-Year) (%)": {
        "components": ["Total Revenue"],
        "operation": "CAGR",
        "years": 5,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Total Revenue"
    },
    "Net Income CAGR (5-Year) (%)": {
        "components": ["Net Income"],
        "operation": "CAGR",
        "years": 5,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Net Income"
    },
    "Total Assets CAGR (5-Year) (%)": {
        "components": ["Total Assets"],
        "operation": "CAGR",
        "years": 5,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Total Assets"
    },
    "Total Common Shares Outstanding CAGR (5-Year) (%)": {
        "components": ["Total Common Shares Outstanding"],
        "operation": "CAGR",
        "years": 5,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Total Common Shares Outstanding"
    },
    
    #CAGR Section: ============================
    
    "Cash From Operations CAGR (3-Year) (%)": {
        "components": ["Cash From Operations"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Cash From Operations"
    },
    "Earnings Per Share (EPS) CAGR (3-Year) (%)": {
        "components": ["Earnings Per Share (EPS)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for EPS"
    },
    "Gross Profit Margin (%) CAGR (3-Year) (%)": {
        "components": ["Gross Profit Margin (%)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Gross Profit Margin"
    },
    "Operating Profit Margin (%) CAGR (3-Year) (%)": {
        "components": ["Operating Profit Margin (%)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Operating Profit Margin"
    },
    "Return On Assets (%) CAGR (3-Year) (%)": {
        "components": ["Return On Assets"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Return On Assets"
    },
    "Return On Equity (%) CAGR (3-Year)": {
        "components": ["Return On Equity"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Return On Equity"
    },
    "Value Spread (Heavy) (%) CAGR (3-Year)": {
        "components": ["Value Spread (Heavy) (%)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Value Spread (Heavy)"
    },
    "Value Spread (Light) (%) CAGR (3-Year)": {
        "components": ["Value Spread (Light) (%)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Value Spread (Light)"
    },
    "Market Cap (%) CAGR (3-Year)": {
        "components": ["Market Cap"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Market Cap"
    },
    "Enterprise Value (%) CAGR (3-Year)": {
        "components": ["Enterprise Value"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Enterprise Value"
    },
    "Tangible Book Value Per Share (%) CAGR (3-Year)": {
        "components": ["Tangible Book Value Per Share"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Tangible Book Value Per Share"
    },
    "P/E Ratio (Calculated) (%) CAGR (3-Year)": {
        "components": ["P/E Ratio (Calculated)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for P/E Ratio (Calculated)"
    },
    "Price To Book (P/B Ratio) (%) CAGR (3-Year)": {
        "components": ["Price To Book (P/B Ratio)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Price To Book (P/B Ratio)"
    },
    "EV / EBITDA (%) CAGR (3-Year)": {
        "components": ["EV / EBITDA"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for EV / EBITDA"
    },
    "EV / Share (%) CAGR (3-Year)": {
        "components": ["EV / Share"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for EV / Share"
    },
    "Total Revenue (%) CAGR (3-Year)": {
        "components": ["Total Revenue"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Total Revenue"
    },
    "Total Revenue CAGR (3-Year)": {
        "components": ["Total Revenue"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 1,
        "description": "3-Year Compound Annual Growth Rate for Total Revenue"
    },
    "Net Income (%) CAGR (3-Year)": {
        "components": ["Net Income"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Net Income"
    },
    "Total Assets (%) CAGR (3-Year)": {
        "components": ["Total Assets"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Total Assets"
    },
    "Cash From Operations (%) CAGR (3-Year)": {
        "components": ["Cash From Operations"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Operating Profit Margin"
    },
    "Shareholder/Stockholder Equity (%) CAGR (3-Year)": {
        "components": ["Shareholder/Stockholder Equity"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Shareholder/Stockholder Equity"
    },
    "Adjusted Debt to Equity Ratio (%) CAGR (3-Year)": {
        "components": ["Adjusted Debt to Equity Ratio"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Adjusted Debt to Equity Ratio"
    },
    "Adjusted Debt to Asset Ratio (%) CAGR (3-Year)": {
        "components": ["Adjusted Debt to Asset Ratio"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Adjusted Debt to Asset Ratio"
    },
    "Quick Ratio 'Liquid Ratio' (%) CAGR (3-Year)": {
        "components": ["Quick Ratio 'Liquid Ratio'"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Quick Ratio 'Liquid Ratio'"
    },
    "Interest Coverage Ratio (%) CAGR (3-Year)": {
        "components": ["Interest Coverage Ratio"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Interest Coverage Ratio"
    },
    "Tax Retention Ratio (%) CAGR (3-Year)": {
        "components": ["Tax Retention Ratio"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Tax Retention Ratio"
    },
    "Asset Turnover Ratio (%) CAGR (3-Year)": {
        "components": ["Asset Turnover Ratio"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Asset Turnover Ratio"
    },
    "Net Profit Margin (%) CAGR (3-Year) (%)": {
        "components": ["Net Profit Margin (%)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Net Profit Margin"
    },
    "ROIC (Heavy) CAGR (3-Year) (%)": {
        "components": ["ROIC (Heavy)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for ROIC (Heavy)"
    },
    "ROIC (Light) CAGR (3-Year) (%)": {
        "components": ["ROIC (Light)"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for ROIC (Light)"
    },
    "Interest Bearing Debt CAGR (3-Year) (%)": {
        "components": ["Interest Bearing Debt"],
        "operation": "CAGR",
        "years": 3,
        "multiplier": 100,
        "description": "5-Year Compound Annual Growth Rate for Interest Bearing Debt"
    },
    "1 + Revenue CAGR": {
        "components": ["Total Revenue CAGR (3-Year)",1],
        "operation": "+",
        "description": "Total Revenue Plus 1"
    },
    "(1 + Revenue CAGR)^3": {
        "components": ["1 + Revenue CAGR","1 + Revenue CAGR","1 + Revenue CAGR"],
        "operation": "*",
        "description": "(1 + Revenue CAGR)^3"
    },
    "Revenue 3-Year Forward": {
        "components": ["(1 + Revenue CAGR)^3","Total Revenue"],
        "operation": "*",
        "description": "One Year Forward Revenue Lookout"
    },
    "Free Cash Flow Margin": {
        "components": ["Free Cash Flow","Total Revenue"],
        "operation": "/",
        "description": "FCF / Total Revenue"
    },
    "FCF Forward (3Y)": {
        "components": ["Revenue 3-Year Forward","Free Cash Flow Margin"],
        "operation": "*",
        "description": "Forward Looking Revenue"
    },
    "EV / FCF Forward (3Y)": {
        "components": ["Enterprise Value","FCF Forward (3Y)"],
        "operation": "/",
        "description": "EV / FCF Forward (3Y)"
    },
    "Gross Profit / EV": {
        "components": ["Gross Profit","Enterprise Value"],
        "operation": "/",
        "description": "Gross Profit / EV"
    },
    "Price / Sales": {
        "components": ["Market Cap","Total Revenue"],
        "operation": "/",
        "description": "Markat Cap / Total Revenue"
    },
    
    
    #=====================================
    
    
    "PEG Ratio (Calculated)": {
        "components": ["P/E Ratio (Calculated)", "Earnings Per Share (EPS) CAGR (3-Year) (%)"],
        "operation": "/",
        "description": "P/E Ratio / Earnings Per Share (EPS) CAGR (3-Year) (%)"
    },
    # --- Solvency Ratios (Level 3) ---
    "Quick Ratio 'Liquid Ratio'": {
        "components": ["Cash and Cash Equivalent + Short Term Investments + Accounts Recievable","Current Liabilities"],
        "operation": "/",
        "description": "(Cash and Cash Equivalent + Short Term Investments + Accounts Recievable) / Current Liabilities"
    },
    "Adjusted Debt to Equity Ratio": {
        "components": ["Net Debt","Shareholder/Stockholder Equity"],
        "operation": "/",
        "description": "Net Debt / Shareholder/Stockholder Equity"
    },
    "Adjusted Debt to Asset Ratio": {
        "components": ["Net Debt","Total Assets"],
        "operation": "/",
        "description": "Net Debt / Total Assets"
    },
    "Interest Coverage Ratio": { #MOST IMPORTANT
        "components": ["EBITDA","Interest Expense Consolidated"],
        "operation": "/",
        "description": "EBITDA / Interest Expense Consolidated"
    },
    
} # <-- END of CALCULATED_EQUATIONS

# =========================================================================
# === EXPORT ORDER LISTS (CALCULATED_EXPORT_ORDER & RAW_EXPORT_ORDER - MODIFIED)
# =========================================================================

# --- Python-Calculated Metrics Export Order ---
CALCULATED_EXPORT_ORDER = [
    
    # --- PROFITABILITY MARGINS ---
    "Gross Profit Margin (%)",
    "Operating Profit Margin (%)",
    "Net Profit Margin (%)",
    
    # --- ROIC/ROA/ROE CHAIN ---
    "ROIC (Heavy)",
    "Return On Assets",
    "Return On Equity",
    "ROIC (Light)",
    
    # --- SPREAD ANALYSIS ---
    "WACC (%)",
    "Value Spread (Heavy) (%)",
    "Value Spread (Light) (%)",
    
    # --- VALUATION MULTIPLES ---
    "Closing Price (USD)",
    "Market Cap",
    "Enterprise Value",
    "Total Common Shares Outstanding",
    "Earnings Per Share (EPS)",
    "Tangible Book Value Per Share",
    "P/E Ratio (Calculated)",
    "Price To Book (P/B Ratio)",
    "EV / EBITDA",
    "EV / Share",
    "PEG Ratio (Calculated)", # <-- KEPT here as requested
    #New
    "Total Revenue",
    "Net Income",
    "Total Assets",
    "Total Common Shares Outstanding",
    "Cash From Operations",
    "Free Cash Flow",
    # --- SOLVENCY RATIOS ---
    "Adjusted Debt to Equity Ratio",
    "Adjusted Debt to Asset Ratio",
    "Quick Ratio 'Liquid Ratio'",
    "Interest Coverage Ratio",
    
    # --- INTERMEDIATE CALCULATIONS (For Modeling/Debugging) ---
    "Net Operating Profit After Taxes (Heavy)",
    "Net Operating Profit After Taxes (Light)",
    "Embedded Lease Interest", # NEW: Added for debugging the NOPAT calc
    "Invested Capital (Heavy)",
    "Invested Capital (Light)",
    "Net Working Capital (Heavy)",
    "Tax Rate",
    "Tax Retention Ratio",
    "Tax Burden",
    "Interest Burden",
    "EBT",
    "EBIT", # NEW: Added EBIT to the export list
    "EBITDA", # NEW: Added EBITDA to the export list
    "Book Value", # NEW: Added Book Value
    "Tangible Book Value", # NEW: Added Tangible Book Value
    "Interest Bearing Debt", # NEW: Added Interest Bearing Debt
    "Net Debt", # NEW: Added Net Debt
    "Operating Margin",
    "Asset Turnover Ratio",
    "Financial Leverage",
    "Interest Expense Consolidated",
    "Average Total Assets",
    "Average Shareholder Equity",
    
    # --- LIGHT ROIC CHAIN INPUTS ---
    "Operating Cash",
    "Excess Cash",
    "Excess Cash + Short Term Investments",
    
    # --- START EXPORT ORDER MODIFICATION ---
    "Operating Current Assets",
    #---- Valuation Screen Data------
    "EV / FCF Forward (3Y)",
    "Gross Profit / EV",
    "Price / Sales",
    "Revenue 3-Year Forward",
    "FCF Forward (3Y)",
    # --- END EXPORT ORDER MODIFICATION ---
    
    "Non Interest Bearing Current Liabilitites",
    "Net Working Capital (Light)",
    "Interest Expense Consolidated * Tax Burden",
    "Net Income + Interest Expense Consolidated * Tax Burden",
    "Tax Burden * Interest Burden",
    "Tax Burden * Interest Burden * Operating Margin",
    "Tax Burden * Interest Burden * Operating Margin * Asset Turnover Ratio",
    # Intermediate components moved from pre-calc must also be added here if needed for debugging
    "Accounts Recievable + Inventory + Prepaid Expenses",
    "Notes Payable + Current Bank Overdraft + Commercial Papers",
    "Net Intangible Assets + Net PP&E + Goodwill",
    "Net Intangible Assets + Net PP&E + Goodwill + Operating Lease ROU Asset",
    "Cash and Cash Equivalent + Short Term Investments + Accounts Recievable",
    
    "Interest Bearing Debt",
    "Shareholder/Stockholder Equity",
    
    #CAGR VALUES
    
    "Interest Bearing Debt CAGR (3-Year) (%)",
    "ROIC (Light) CAGR (3-Year) (%)",
    "ROIC (Heavy) CAGR (3-Year) (%)",
    "Net Profit Margin (%) CAGR (3-Year) (%)",
    "Asset Turnover Ratio (%) CAGR (3-Year)",
    "Tax Retention Ratio (%) CAGR (3-Year)",
    "Interest Coverage Ratio (%) CAGR (3-Year)",
    "Quick Ratio 'Liquid Ratio' (%) CAGR (3-Year)",
    "Adjusted Debt to Asset Ratio (%) CAGR (3-Year)",
    "Adjusted Debt to Equity Ratio (%) CAGR (3-Year)",
    "Shareholder/Stockholder Equity (%) CAGR (3-Year)",
    "Cash From Operations (%) CAGR (3-Year)",
    "Earnings Per Share (EPS) CAGR (3-Year) (%)",
    "Gross Profit Margin (%) CAGR (3-Year) (%)",
    "Operating Profit Margin (%) CAGR (3-Year) (%)",
    "Return On Assets (%) CAGR (3-Year) (%)",
    "Return On Equity (%) CAGR (3-Year)",
    "Value Spread (Heavy) (%) CAGR (3-Year)",
    "Value Spread (Light) (%) CAGR (3-Year)",
    "Market Cap (%) CAGR (3-Year)",
    "Enterprise Value (%) CAGR (3-Year)",
    "Tangible Book Value Per Share (%) CAGR (3-Year)",
    "P/E Ratio (Calculated) (%) CAGR (3-Year)",
    "Price To Book (P/B Ratio) (%) CAGR (3-Year)",
    "EV / EBITDA (%) CAGR (3-Year)",
    "EV / Share (%) CAGR (3-Year)",
    "Total Revenue (%) CAGR (3-Year)",
    "Net Income (%) CAGR (3-Year)",
    "Total Assets (%) CAGR (3-Year)",
    "(1 + Revenue CAGR)^3",
    "1 + Revenue CAGR",
    "Revenue 3-Year Forward",
    "Free Cash Flow Margin",
    "FCF Forward (3Y)",
    
    
    
    
    
    
    
    
    
    
    
]


# --- RAW DATA EXPORT ORDER ---
RAW_EXPORT_ORDER = [
    # --- Income Statement ---
    "Total Revenue",
    "Cost of Goods Sold",
    "Gross Profit",
    "Operating Expense (EBIT)",
    "Operating Income",
    "Depreciation, Depletion, and Amortization",
    "Net Income",
    "Income Before Tax",
    "Income Tax Expense",
    "Interest Expense",
    "Interest Expense Non-Operating",
    # Helper Calcs (now handled in CALCULATED_EQUATIONS, but may exist as raw if scraped)
    "EBITDA",
    "EBIT",
    
    # --- Balance Sheet (Assets) ---
    "Total Assets",
    "Current Assets",
    "Cash & Cash Equivelance",
    "Short Term Investments",
    "Accounts Receivable",
    "Inventory",
    "Prepaid Expenses", # Helper Calc
    "Net PP&E",
    "Goodwill",
    "Net Intangible Assets",
    "Operating Lease ROU Asset", # NEW: Added ROU Asset
    
    # --- Balance Sheet (Liabilities & Equity) ---
    "Total Liabilities",
    "Current Liabilities",
    "Notes Payable",
    "Commercial Papers",
    "Current Bank Overdraft",
    "Short Term Debt (STD)",
    "Long Term Debt (LTD)",
    "Convertible Senior Notes",
    "Operating Lease Liability",
    "Weighted Average Lease Discount Rate", # NEW: Added WADR
    # Helper Calcs (now handled in CALCULATED_EQUATIONS)
    "Interest Bearing Debt",
    "Net Debt",
    "Preferred Stock",
    "Preferred Stock Issued",
    "Shareholder/Stockholder Equity",
    "Minority Interest",
    # Helper Calcs (now handled in CALCULATED_EQUATIONS)
    "Book Value",
    "Tangible Book Value",
    "Total Common Shares Outstanding",
    
    # --- Cash Flow Statement (Operating & Investing) ---
    "Cash From Operations",
    "SegmentExpenditureAdditionToLongLivedAssets",
    "Payments To Acquire PP&E",
    "Payments To Acquire Intangible Assets",
    "Payments To Acquire Investments",
    "Pay. Acquire Other Investment",
    "Payments To Acquire Businesses Net Of Cash Acquired",
    
    # --- Cash Flow Statement (Financing) ---
    "Proceed From Issuance of New Stock / Equity",
    "Proceeds From Issuance Of Debt",
    "Proceed From Long Term Debt",
    "Proceed From Short Term Debt",
    "Payments For Repurchase of Stock",
    "Dividend Common Stock Cash",
    "Repayment Of Debt",
    "Repayment Of Long Term Debt",
    "Repayment Of Short Term Debt",
    "Finance Lease Principal Payments",
    
    # --- Compound/Helper Tags (For Calculation Purposes Only) ---
    "Accounts Recievable + Inventory + Prepaid Expenses",
    "Notes Payable + Current Bank Overdraft + Commercial Papers",
    "Net Intangible Assets + Net PP&E + Goodwill",
    "Net Intangible Assets + Net PP&E + Goodwill + Operating Lease Liability",
    "Cash and Cash Equivalent + Short Term Investments + Accounts Recievable",
]



#====================Is Capital Shifting=======================
SEMICONDUCTORS = [
    "NVDA", "AMD", "INTC", "MU", "MRVL", "QCOM", "TXN", "NXPI", "ON", "ADI", "MCHP", "MPWR", "SWKS", "QRVO", "LSCC", "SYNA", "CRUS", "POWI", "AOSL"
]

SEMICONDUCTORS_CONFIRMERS = [
    "SITM", "SLAB", "NVEC", "INDI", "MXL", "CRDO", "MX", "DIOD"
]

APPLICATION_SOFTWARE = [
    "ADBE", "ADSK", "CRM", "NOW", "DDOG", "SNOW", "MDB", "WDAY", "TEAM", "INTU", "HUBS", "GTLB", "DOCU", "ESTC", "CFLT", "FICO", "BSY", "MANH", "TYL", "QTWO", "PTC", "GWRE"
]

APPLICATION_SOFTWARE_CONFIRMERS = [
    "ACIW", "AGYS", "AI", "AIOT", "AKAM", "ALKT", "ALRM", "AMPL", "APPF", "APPN", "APPS", "BASE", "BILL", "BL", "BLKB", "BLND", "BOX", "BRZE", "CXM", "CWAN", "DBX", "DOMO", "DT", "EXFY", "EXOD", "FIVN", "FROG", "FRSH", "GDDY", "IOT", "INFA", "INTA", "JAMF", "LAW", "LIF", "MIR", "MLNK", "NABL", "NCNO", "NTNX", "ONTF", "PAR", "PATH", "PD", "PEGA", "PENG", "PLTR", "PRO", "PRGS", "RAMP", "RDVT", "RNG", "SEMR", "SOUN", "SPT", "SPSC", "TDC", "VERX", "WEAV", "WK", "YEXT", "ZETA", "ZM", "APP", "ASAN", "AVPT"
]

SEMICONDUCTORS_EQUIPTMENT = [
    "AMAT", "LRCX", "KLAC", "ASML", "ENTG", "ACLS"
    ]

SEMICONDUCTORS_EQUIPTMENT_CONFIRMERS = [
    "ONTO", "FORM", "ACMR", "PLAB", "UCTT", "COHU", "PDFS", "VECO", "ICHR", "KLIC", "ASYS", "RTEC", "CAMT"
]

HARDWARE_AND_STORAGE = [
    "DELL", "HPQ", "HPE", "NTAP", "STX", "WDC", "SMCI", "GLW", "CIEN", "COMM", "TEL", "SNX", "SANM", "JBL", "FLEX", "PLXS", "TTMI", "KEYS", "TDY", "VSH", "AVT", "APH"
]

HARDWARE_AND_STORAGE_CONFIRMATION = [
    "AAOI", "ADTN", "ARLO", "ATEN", "BELFA", "BELFB", "BMI", "CLFD", "CRSR", "DAKT", "ITRI", "KE", "LFUS", "LITE", "LWLG", "MEI", "MVIS", "NTGR", "OSIS", "OUST", "ROG", "SCSC", "VPG", "XRX", "ZBRA"
]

SEARCH_AND_DIGITAL_MEDIA = [
    "TTD", "MGNI", "IAS"
]

SEACHR_AND_DIGITAL_MEDIA_CONFIRMATION = [
    "PERI", "DV"
]

#====================Is Capital Spreading======================


SYSTEM_SOFTWARE = [
    "ORCL", "NOW", "VMW", "NTAP"
]

SYSTEM_SOFTWARE_CONFIRMATION = [
    "AKAM", "DOCN", "FFIV", "TDC", "VRSN", "NTCT"
]

CYBERSECURITY = [
    "CRWD", "PANW", "FTNT", "ZS", "OKTA"
]

CYBERSECURITY_CONFIRMATION = [
    "TENB", "QLYS", "RPD", "S", "GEN", "VRNS", "CYBR"
]

E-COMMERCE_MARKETPLACE = [
    "EBAY", "ETSY", "SHOP", "WISH"
]

MOBILITY_AND_DELIVERY_PLATFORM = [
    "UBER", "LYFT", "DASH"
]

INTERACTIVE_HOME_ENTERTAINMENT = [
    "EA", "TTWO", "RBLX"
]

INTERACTIVE_HOME_ENTERTAINMENT_COMPARISON = [
    "PLTK", "SKLZ"
]


#====================Is this Late / End-Cycle==================


IT_CONSULTING_AND_SERVICES = [
    "ACN", "CTSH", "DXC", "EPAM", "SAIC", "CACI", "G"
]

IT_CONSULTING_AND_SERVICES_CONFIRMATION = [
    "ASGN", "KD", "RXT", "UIS", "PRFT", "BGSF", "CNDT", "PSN"
]

INTEGRATED_TELECOM = [
    "VZ", "T", "TMUS", "LUMN", "CABO", "TDS"
]

INTEGRATED_TELECOM_CONFIRMATION = [
    "FYBR", "OPTU", "WOW", "SHEN", "CCOI", "ATNI", "GOGO"
]

TRAVEL_AND_ACCOMODATION = [
    "ABNB", "BKNG", "EXPE"
]


#===========Regimine Groups identifier========


REGIME_GROUPS = {
#====================Is Capital Shifting=======================

    "Semiconductors": {
        "core": SEMICONDUCTORS,
        "confirmers": SEMICONDUCTORS_CONFIRMERS
    },
    "Application Software": {
        "core": APPLICATION_SOFTWARE,
        "confirmers": APPLICATION_SOFTWARE_CONFIRMERS
    },
    "Semiconductor Equiptment": {
        "core": SEMICONDUCTORS_EQUIPTMENT,
        "confirmers": SEMICONDUCTORS_EQUIPTMENT_CONFIRMERS
    },
    "Hardware and Storage": {
        "core": HARDWARE_AND_STORAGE,
        "confirmers": HARDWARE_AND_STORAGE_CONFIRMATION
    },
    "Search and Digital Media": {
        "core": SEARCH_AND_DIGITAL MEDIA,
        "confirmers": SEARCH_&_DIGITAL_MEDIA_CONFIRMATION
    },
#====================Is Capital Spreading======================

    "System Software": {
        "core": SYSTEM_SOFTWARE,
        "confirmers": SYSTEM_SOFTWARE_CONFIRMATION
    },
    "Cybersecurity": {
        "core": CYBERSECURITY,
        "confirmers": CYBERSECURITY_CONFIRMATION
    },
    "E-Commerce Marketplace": { #core only
        "core": E-COMMERCE_MARKETPLACE
    },
    "Mobility and Delivery Platform": { #core only
        "core": MOBILITY_AND_DELIVERY_PLATFORM
    },
    "Interactive Home Entertainment": {
        "core": INTERACTIVE_HOME_ENTERTAINMENT,
        "confirmers": INTERACTIVE_HOME_ENTERTAINMENT_COMPARISON
    },
#====================Is this Late / End-Cycle==================

    "IT Counsulting and Services": {
        "core": IT_CONSULTING_AND_SERVICES,
        "confirmers": IT_CONSULTING_&_SERVICES_CONFIRMATION
    },
    "Integrated Telecom": {
        "core": INTEGRATED_TELECOM,
        "confirmers": INTEGRATED_TELECOM_CONFIRMATION
    },
    "Travel and Accomodation": {
        "core": TRAVEL_AND_ACCOMODATION
    }
}



# =========================================================================
# === HELPER FUNCTIONS (MAIN FETCH LOGIC UNCHANGED, BUT SUPPORT FUNCTIONS ARE FIXED) ===
# =========================================================================
# REGIME Helper Function

def fetch_polygon_daily_prices(
    ticker: str,
    start_date: str,
    end_date: str,
    api_key: str,
    sleep_sec: float = 0.25
) -> pd.DataFrame:
    """
    Fetch daily OHLCV data from Polygon.io for one ticker.
    """

    url = (
        f"https://api.polygon.io/v2/aggs/ticker/{ticker}/range/1/day/"
        f"{start_date}/{end_date}"
    )

    params = {
        "adjusted": "true",
        "sort": "asc",
        "limit": 50000,
        "apiKey": api_key,
    }

    r = requests.get(url, params=params)
    r.raise_for_status()
    data = r.json()

    if "results" not in data or not data["results"]:
        return pd.DataFrame()

    df = pd.DataFrame(data["results"])

    df["Date"] = pd.to_datetime(df["t"], unit="ms")
    df.set_index("Date", inplace=True)

    df = df.rename(columns={
        "o": "Open",
        "h": "High",
        "l": "Low",
        "c": "Close",
        "v": "Volume",
    })

    time.sleep(sleep_sec)

    return df[["Open", "High", "Low", "Close", "Volume"]]
    

#Second REGIME helper def
def flatten_ticker_groups(regime_groups: dict) -> list:
    tickers = set()
    for group in regime_groups.values():
        tickers.update(group.get("core", []))
        tickers.update(group.get("confirmers", []))
    return sorted(tickers)


# Third REGIME helper def
def fetch_all_prices(
    tickers: list,
    start_date: str,
    end_date: str
) -> dict:
    price_data = {}

    for ticker in tickers:
        try:
            df = fetch_polygon_daily_prices(
                ticker=ticker,
                start_date=start_date,
                end_date=end_date,
                api_key=POLYGON_API_KEY
            )
            if not df.empty:
                price_data[ticker] = df
        except Exception as e:
            print(f"[WARN] Failed {ticker}: {e}")

    return price_data

#========================Price Data REGIME Starting Equations===============
#Simple Moving Average
def compute_sma(series, window):
    return series.rolling(window).mean()

#Percent Difference from SMA
def percent_from_sma(price_series, sma_series):
    return (price_series - sma_series) / sma_series

#Low Detection
def is_new_low(price_series, lookback):
    rolling_min = price_series.rolling(lookback).min()
    return price_series == rolling_min

#Three come together for this
def build_price_features(df):
    """
    df must contain a 'close' column indexed by date
    """

    df = df.copy()

    # Moving Averages
    df["SMA_20"] = compute_sma(df["close"], 20)
    df["SMA_50"] = compute_sma(df["close"], 50)

    # Percent Distance
    df["Pct_From_SMA_20"] = percent_from_sma(df["close"], df["SMA_20"])
    df["Pct_From_SMA_50"] = percent_from_sma(df["close"], df["SMA_50"])

    # New Lows
    df["New_Low_20D"] = is_new_low(df["close"], 20)
    df["New_Low_50D"] = is_new_low(df["close"], 50)

    return df

#========================================================================

#Valid trading days from Polygon
def get_trading_days(start_date, end_date):
    """
    Returns a list of trading dates (YYYY-MM-DD) between start_date and end_date.
    """
    return pd.date_range(start=start_date, end=end_date, freq="B").strftime("%Y-%m-%d").tolist()

#Detects missing dates
def detect_missing_trading_days(history_df, today):
    """
    Returns a list of trading dates that are missing from history.
    """
    if history_df.empty:
        return []

    history_df["Date"] = pd.to_datetime(history_df["Date"])
    last_date = history_df["Date"].max().date()
    
    all_days = pd.date_range(start=last_date + pd.Timedelta(days=1),
                             end=pd.to_datetime(today),
                             freq="B")

    return [d.strftime("%Y-%m-%d") for d in all_days]

#Backfill engine
def backfill_missing_days(
    missing_dates,
    compute_snapshot_fn,
    history_df
):
    """
    Fills missing trading days sequentially.
    """
    rows = []

    for date_str in missing_dates:
        snapshot = compute_snapshot_fn(date_str)

        if snapshot is not None:
            snapshot["Date"] = date_str
            rows.append(snapshot)

    if rows:
        history_df = pd.concat([history_df, pd.DataFrame(rows)], ignore_index=True)
        history_df = history_df.sort_values("Date").reset_index(drop=True)

    return history_df

#=======SubIndustry Snapshot=================

def compute_subindustry_snapshot(
    date_str: str,
    subindustry_name: str,
    subindustry_group: dict,
    price_data: dict
):
    """
    Builds one daily snapshot row for a sub-industry.

    Parameters
    ----------
    date_str : str (YYYY-MM-DD)
    subindustry_name : str
    subindustry_group : dict with keys ["core", "confirmers"]
    price_data : dict[ticker -> DataFrame] with OHLC data

    Returns
    -------
    dict or None
    """

    tickers = subindustry_group.get("core", []) + subindustry_group.get("confirmers", [])

    rows = []

    for ticker in tickers:
        df = price_data.get(ticker)

        if df is None or df.empty:
            continue

        if date_str not in df.index:
            continue

        df_feat = build_price_features(df)

        row = df_feat.loc[date_str]

        if pd.isna(row["SMA_50"]):
            continue

        rows.append({
            "Pct_From_SMA_20": row["Pct_From_SMA_20"],
            "Pct_From_SMA_50": row["Pct_From_SMA_50"],
            "New_Low_20D": int(row["New_Low_20D"]),
            "New_Low_50D": int(row["New_Low_50D"])
        })

    if not rows:
        return None

    agg = pd.DataFrame(rows)

    snapshot = {
        "Date": date_str,
        "SubIndustry": subindustry_name,

        # ===== TREND DIRECTION (EQUATION 1) =====
        "Median_Pct_From_SMA_20": agg["Pct_From_SMA_20"].median(),
        "Median_Pct_From_SMA_50": agg["Pct_From_SMA_50"].median(),

        # ===== SELLING PRESSURE EXHAUSTION (EQUATION 3) =====
        "New_Low_Ratio_20D": agg["New_Low_20D"].mean(),
        "New_Low_Ratio_50D": agg["New_Low_50D"].mean(),

        # ===== BREADTH RECOVERY (EQUATION 5) =====
        "Pct_Above_SMA_20": (agg["Pct_From_SMA_20"] > 0).mean(),
        "Pct_Above_SMA_50": (agg["Pct_From_SMA_50"] > 0).mean(),

        # ===== META =====
        "Stock_Count": len(agg)
    }

    return snapshot

# --- Helper Function 1: Get CIK from Ticker Symbol (Unchanged) ---
def get_cik_from_ticker(ticker_symbol):
    """ Fetches the CIK (Central Index Key) for a given stock ticker symbol. """
    print(f"Finding CIK for ticker: {ticker_symbol}...")
    try:
        url = "https://www.sec.gov/files/company_tickers.json"
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()
        tickers_data = response.json()
        for entry in tickers_data.values():
            if entry['ticker'] == ticker_symbol.upper():
                cik = str(entry['cik_str']).zfill(10)
                print(f"CIK for {ticker_symbol} found: {cik}")
                return cik
        print(f"CIK for {ticker_symbol} not found in SEC database.")
        return None
    except requests.exceptions.RequestException as e:
        print(f"Error fetching CIK data: {e}")
        return None

# --- Helper Function 2: Get Current Share Price & Historical Prices (Unchanged) ---
def get_historical_share_prices(ticker_symbol):
    """
    Fetches the current share price and historical prices for the end of each of the
    last 5 full fiscal years (Dec 31st) from Yahoo Finance.
    """
    try:
        # Fetch data for the ticker
        ticker = yf.Ticker(ticker_symbol)
        
        # Get historical data for the last 6 years (to cover 5 full fiscal years)
        # We fetch daily data to ensure we find a closing price on or near Dec 31st
        current_year = date.today().year
        # Start date 6 years ago to capture the 5th prior year's closing price
        start_year = current_year - 6
        start_date = date(start_year, 1, 1).strftime('%Y-%m-%d')
        end_date = date.today().strftime('%Y-%m-%d')
        history = ticker.history(start=start_date, end=end_date, interval='1d')
        
        if history.empty:
            print(f"Warning: No historical price data found for {ticker_symbol}.")
            return None

        historical_prices = {}
        # Iterate over the last 5 full fiscal years
        for i in range(1, 6):
            target_year = current_year - i
            # Assume fiscal year ends on Dec 31st for simplicity (e.g., Dec 31st)
            target_date = date(target_year, 12, 31)
            
            # Find the closest trading day on or before the target date
            closest_price_day = history.loc[history.index <= str(target_date)].tail(1)
            
            if not closest_price_day.empty:
                # Use the 'Close' price for that day
                historical_prices[str(target_year)] = closest_price_day['Close'].iloc[0]
            # else:
                # print(f"Warning: No closing price found for year {target_year}.")

        # Convert to Pandas Series, sorted by year descending
        prices_series = pd.Series(historical_prices).sort_index(ascending=False)
        prices_series.name = "Closing Price (USD)"
        print("Historical prices fetched successfully.")
        return prices_series
    except Exception as e:
        print(f"An error occurred while fetching historical share prices: {e}")
        return None

# MODIFIED HELPER FUNCTION to correctly find non-monetary (ratio) footnote data (Unchanged)
def get_annual_data(sec_facts, label, expected_unit='USD'):
    """
    Extracts annual (10-K) data points for a specific SEC label.
    Prioritizes USD unit unless explicitly looking for 'shares'.
    Checks for 'ratio' or 'pure' as a final fallback for footnote values like WADR.
    """
    data = {}
    if label in sec_facts:
        details = sec_facts[label]
        
        # Determine unit key (either USD, shares, or the first available if not specified)
        unit_key = details['units'].get(expected_unit)
        if not unit_key and expected_unit == 'USD':
            # Check for shares if USD not found
            unit_key = details['units'].get('shares')
            
        # --- NEW ADDITION: Check for non-monetary units (ratio/pure) if USD/shares not found ---
        if not unit_key:
            unit_key = details['units'].get('ratio')
        if not unit_key:
            unit_key = details['units'].get('pure')
        
        if unit_key:
            for fact in unit_key:
                # Only interested in annual (10-K) data
                if fact.get('form') == '10-K':
                    # Extract the year from the end date (e.g., "2023-12-31" -> "2023")
                    year = fact.get('end', '').split('-')[0]
                    # Only store if the year is a valid 4-digit number
                    if year and len(year) == 4 and year.isdigit():
                        data[year] = fact.get('val')
    
    return data

# --- NEW HELPER FUNCTION: Smartly check for stock split factors (Unchanged) ---
def _check_split_ratio(ratio):
    """
    Checks if a calculated ratio (new_shares/old_shares) corresponds to a common
    stock split (e.g., 2:1, 3:1, 4:1). Returns the integer factor or 1.0.
    """
    # Common split factors
    common_splits = [2, 3, 4, 5, 10, 100]
    
    for split in common_splits:
        # Check if the ratio is very close to a common integer split
        if abs(ratio - split) < 0.05: # Use a small tolerance
            return float(split)
            
    # Also check for common reverse splits (e.g., 1-for-2 = 0.5)
    reverse_splits = [0.5, 0.333, 0.25, 0.2, 0.1, 0.01]
    for split in reverse_splits:
        if abs(ratio - split) < 0.05:
            return float(ratio) # Apply the inverse splits (e.g., 1-for-2 = 0.5) - this should not happen here since ratio >= 1.5

    # If no common split is found, return the original ratio or 1.0 (no split)
    return ratio if ratio > 1.0 else 1.0

# --- NEW FUNCTION: Apply Historical Stock Split Adjustment (Unchanged) ---
def apply_stock_split_adjustment(financial_data):
    """
    Detects and adjusts historical share counts and closing prices based on stock splits.
    This is necessary because SEC data is not retrospectively adjusted.
    """
    # 1. Get the raw shares outstanding data
    shares_data = financial_data.get("Total Common Shares Outstanding", {})
    if not shares_data:
        print("No shares outstanding data found for split adjustment.")
        return financial_data

    # 2. Detect Splits and Factors
    split_factors = {} # {year_split_took_effect: split_factor}
    
    # Sort years ascending to compare prior year to current year
    sorted_years = sorted([int(y) for y in shares_data.keys()])

    for i in range(1, len(sorted_years)):
        current_year = sorted_years[i]
        prior_year = sorted_years[i-1]
        
        # Keys as strings for dictionary access
        current_year_str = str(current_year)
        prior_year_str = str(prior_year)
        
        current_shares = shares_data.get(current_year_str)
        prior_shares = shares_data.get(prior_year_str)

        # Ensure we have valid, non-zero numeric data
        try:
            current_shares = float(current_shares)
            prior_shares = float(prior_shares)
        except:
            continue
            
        if prior_shares == 0:
            continue
            
        ratio = current_shares / prior_shares
        
        # Check for a forward split (where shares *increase* dramatically)
        # Use a high threshold like 1.5 (e.g., 2-for-1 split = 2.0)
        if ratio >= 1.5:
            split_factor = _check_split_ratio(ratio)
            # A split of N:M means shares *increase* by N/M.
            # We want the factor to multiply *old* data to make it comparable to *new* data.
            # Example: 2-for-1 split. 2023 shares = 100. 2022 shares = 50. Ratio = 2.0.
            # We want to multiply 2022 shares by 2.0, so the split factor is the ratio.
            if split_factor > 1.0:
                 # Store the factor that makes old data equal to new data.
                split_factors[current_year_str] = split_factor

    # 3. Calculate Cumulative Adjustment Factors
    # This factor will be applied to ALL historical data BEFORE the split year.
    adjustment_factors_by_year = {str(year): 1.0 for year in sorted_years}
    cumulative_factor = 1.0
    
    # Iterate years in reverse chronological order (most recent split first)
    for year in sorted(split_factors.keys(), key=int, reverse=True):
        split_factor = split_factors[year]
        # Multiply the cumulative factor by the new split factor
        cumulative_factor *= split_factor
        
        # Apply the current cumulative factor to all years *prior* to the split year
        for pre_split_year in [y for y in sorted_years if int(y) < int(year)]:
            # This is the total adjustment needed for data in pre_split_year
            adjustment_factors_by_year[str(pre_split_year)] = cumulative_factor

    # We must apply the last cumulative factor to the oldest year as well.
    # The last factor represents the total multiplier for all years prior to the oldest split year.
    
    # 4. Apply Adjustments to Relevant Metrics
    # Metrics to adjust: Total Common Shares Outstanding, Preferred Stock Issued, Closing Price (USD)
    for metric_name in ["Total Common Shares Outstanding", "Preferred Stock Issued", "Closing Price (USD)"]:
        metric_data = financial_data.get(metric_name, {})
        for year_str in metric_data.keys():
            factor = adjustment_factors_by_year.get(year_str)
            
            if factor is None or factor == 1.0: # Only apply adjustments if needed
                continue
                
            raw_value = metric_data.get(year_str)
            if raw_value is None:
                continue
                
            try:
                raw_value = float(raw_value)
            except:
                continue
                
            # Shares/Preferred Stock Issued: Multiply by factor (e.g., 2021 shares * 2)
            if metric_name in ["Total Common Shares Outstanding", "Preferred Stock Issued"]:
                metric_data[year_str] = raw_value * factor
            # Closing Price (USD): Divide by factor (e.g., 2021 price / 2)
            elif metric_name == "Closing Price (USD)":
                if factor != 0:
                    metric_data[year_str] = raw_value / factor

    return financial_data

# --- NEW HELPER FUNCTION: Process a single fallback item (tag or calculation) (Unchanged) ---
def _process_fallback_item(item, us_gaap_facts, recent_years_list, all_relevant_labels):
    """
    Fetches data for a single SEC tag or performs a calculation based on the item.
    Returns a dictionary of result metadata including the data and year counts.
    """
    current_data = {}
    calc_str = ""
    is_calculation = False
    
    # --- Case 1: Simple Tag Lookup (item is a string) ---
    if isinstance(item, str):
        if item in all_relevant_labels.keys():
            # This should not happen in the fallback loop (it would be a calculated metric)
            pass
        current_data = get_annual_data(us_gaap_facts, item, expected_unit='USD')
        
    # --- Case 2: Multi-Component Calculation (item is a list) ---
    elif isinstance(item, list) and item and item[0] == "CALCULATION:":
        is_calculation = True
        labels = []
        operators = []
        
        # Parse the CALCULATION: structure
        # Structure is ["CALCULATION:", A, OP_1, B, OP_2, C, ...]
        for i, token in enumerate(item[1:]):
            if i % 2 == 0: # Even index (0, 2, 4...) is a label
                labels.append(token)
            else:          # Odd index (1, 3, 5...) is an operator
                operators.append(token)

        num_components = len(labels) # A + B + C -> 3 components
        
        # Perform Calculation if components are valid
        if num_components >= 2:
            component_data = {}
            all_labels_found = True
            common_years = None
            
            # 2a. Fetch all required raw component data
            for label in labels:
                # Do not look up calculated metrics inside the fallback loop, they should be raw tags.
                if label in all_relevant_labels:
                    continue
                data_points = get_annual_data(us_gaap_facts, label, expected_unit='USD')
                if not data_points:
                    all_labels_found = False
                    break
                component_data[label] = data_points
                
                years_set = set(data_points.keys())
                common_years = years_set if common_years is None else common_years.intersection(years_set)

            # 2b. Perform the calculation on common years
            if all_labels_found and common_years:
                for year in common_years:
                    try:
                        # Start with component A (labels[0])
                        calculated_val = float(component_data[labels[0]][year])
                        
                        # Apply operators and components B, C, D
                        for i in range(1, num_components):
                            val_i = float(component_data[labels[i]][year])
                            operator = operators[i-1]
                            
                            if operator == '+':
                                calculated_val += val_i
                            elif operator == '-':
                                calculated_val -= val_i
                            # NOTE: No support for * and / is assumed for SEC calculation fallbacks
                                
                        current_data[year] = calculated_val
                    except Exception:
                        # Catch potential errors during calculation (e.g., non-numeric data)
                        pass
                        
            # Construct the calculation string for printing
            calc_str = f"{labels[0]}"
            for i in range(1, num_components):
                calc_str += f" {operators[i-1]} {labels[i]}"

    # --- Calculate completeness metrics ---
    recent_year_count = sum(1 for year in recent_years_list if year in current_data)
    total_year_count = len(current_data)
    
    return {
        'item': item,
        'data': current_data,
        'is_calculation': is_calculation,
        'recent_year_count': recent_year_count,
        'total_year_count': total_year_count,
        'calc_str': calc_str
    }

# --- MODIFIED: Main Function to FETCH and RETURN DataFrames (Unchanged from previous step) ---
def fetch_financial_data(ticker_symbol):
    """
    Fetches raw SEC financial data, performs necessary calculations, and returns
    DataFrames for raw data and calculated metrics.
    """
    
    # 0. Define Target Lookback Years
    current_year = date.today().year
    TARGET_RECENT_COUNT = 4 # Target must be >= 1
    # Get the last 4 full fiscal years (e.g., for 2024, get 2023, 2022, 2021, 2020)
    recent_years_list = [str(current_year - i) for i in range(1, TARGET_RECENT_COUNT + 1)]
    
    cik = get_cik_from_ticker(ticker_symbol)
    if not cik:
        return None, None, None

    # --- Step 0.5: Fetch Raw SEC Data JSON ---
    print("Fetching raw SEC data JSON...")
    url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
    try:
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()
        data = response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching SEC facts for {ticker_symbol}: {e}")
        return None, None, None

    us_gaap_facts = data.get('facts', {}).get('us-gaap', {})
    financial_data = {} # Stores raw SEC data (keys from RELEVANT_LABELS)

    print(f"Required recent years: {', '.join(recent_years_list)}.")

    # --- Step 1: Fetch Raw SEC Financial Data and Perform Fallback Calculations ---
    print("\n--- Step 1: Fetching Raw SEC Financial Data ---")
    for excel_name, fallbacks in RELEVANT_LABELS.items():
        if not fallbacks:
            continue
            
        all_candidates = []
        for item in fallbacks:
            candidate = _process_fallback_item(item, us_gaap_facts, recent_years_list, RELEVANT_LABELS)
            all_candidates.append(candidate)
            
        found_data = None
        tag_candidates = [c for c in all_candidates if not c['is_calculation']]
        calc_candidates = [c for c in all_candidates if c['is_calculation']]

        # PASS 1: Direct Tags with Full Completeness (4 recent years)
        for candidate in tag_candidates:
            if candidate['recent_year_count'] == TARGET_RECENT_COUNT:
                found_data = candidate
                # print(f"  {excel_name}: Found perfect tag: {candidate['item']}")
                break
                
        # PASS 2: Calculations with Full Completeness (4 recent years)
        if found_data is None:
            for candidate in calc_candidates:
                if candidate['recent_year_count'] == TARGET_RECENT_COUNT:
                    found_data = candidate
                    # print(f"  {excel_name}: Found perfect calculation: {candidate['calc_str']}")
                    break
                    
        # PASS 3: Best Effort (Highest Recent Year Count)
        if found_data is None:
            best_effort_candidate = max(all_candidates, key=lambda c: c['recent_year_count'], default=None)
            if best_effort_candidate and best_effort_candidate['recent_year_count'] > 0:
                found_data = best_effort_candidate
                # if best_effort_candidate['is_calculation']:
                #     print(f"  {excel_name}: Found best effort calculation: {best_effort_candidate['calc_str']} ({best_effort_candidate['recent_year_count']}/{TARGET_RECENT_COUNT})")
                # else:
                #     print(f"  {excel_name}: Found best effort tag: {best_effort_candidate['item']} ({best_effort_candidate['recent_year_count']}/{TARGET_RECENT_COUNT})")
            
        # 4. Store the best result
        if found_data:
            financial_data[excel_name] = found_data['data']
        else:
            # print(f"  {excel_name}: No data found.")
            pass


    # --- Step 2: Inject Historical Share Price ---
    print("\n--- Step 2: Injecting Share Price Data ---")
    price_dict = get_historical_share_prices(ticker_symbol)
    if price_dict is not None:
        financial_data["Closing Price (USD)"] = price_dict.to_dict()
    else:
        print("Warning: Could not fetch historical share prices.")
        
    # === Step 3: STOCK SPLIT ADJUSTMENT ===
    print("\nApplying historical stock split adjustment...")
    financial_data = apply_stock_split_adjustment(financial_data)
        
    # === Step 4: Ensure Data Integrity (Fill missing tags, 0.0 for missing years) ===
    # Get all years found across all collected data
    all_years_set = set()
    for data_dict in financial_data.values():
        all_years_set.update(data_dict.keys())
    all_years_list = sorted(list(all_years_set), key=int, reverse=True) # Sort descending by year
    
    if not all_years_list:
        print(f"No financial data found for {ticker_symbol} across any year.")
        return None, None, None

    # Ensure ALL relevant labels exist as keys in financial_data
    for excel_name in RELEVANT_LABELS.keys():
        if excel_name not in financial_data:
            financial_data[excel_name] = {}
            
    # Fill missing years in the data for this metric with 0.0
    for excel_name in RELEVANT_LABELS.keys():
        current_metric_data = financial_data[excel_name]
        for year in all_years_list:
            if year not in current_metric_data:
                current_metric_data[year] = 0.0
                
    # === Step 5: Convert all values to float to ensure consistency for calculations ===
    for metric_name, data_dict in financial_data.items():
        for year, value in data_dict.items():
            try:
                # Attempt conversion, handling None/NaN/empty strings gracefully
                if value is None or (isinstance(value, str) and not value.strip()):
                    data_dict[year] = np.nan
                else:
                    data_dict[year] = float(value)
            except ValueError:
                data_dict[year] = np.nan # If conversion fails, mark as NaN

    # === Step 6: Calculate Derived Metrics ===
    print("\n--- Step 6: Calculating Derived Metrics and Ratios ---")
    calculated_metrics_df, static_results = calculate_derived_metrics(financial_data, recent_years_list)

    # --- Step 7: Finalize Raw Data DataFrame ---
    raw_df = pd.DataFrame(financial_data).transpose()
    raw_has_data = not raw_df.empty
    
    if raw_has_data:
        # Filter to only keep year columns (which are strings of digits)
        raw_df = raw_df[[col for col in raw_df.columns if str(col).isdigit()]]
        # Reindex to match the desired export order
        raw_df = raw_df.reindex(RAW_EXPORT_ORDER)
        # Sort years descending (most recent first)
        raw_df = raw_df.sort_index(axis=1, ascending=False)
    
    # MODIFIED: Return the DataFrames and static results
    return raw_df, calculated_metrics_df, static_results

# --- New Function: Calculate Derived Metrics (Including CAGR) (FIXED COMPLEX NUMBER ISSUE) ---
def calculate_derived_metrics(financial_data, recent_years_list):
    """
    Calculates time-series metrics (CALCULATED_EQUATIONS) and static values (STATIC_ASSUMPTIONS).
    Returns them as a time-series DataFrame and a dictionary of static name:value pairs.
    """
    time_series_results = {}
    static_results = {}
    
    # Make a copy for intermediate calculations to prevent modifying the raw data
    calculated_data_cache = copy.deepcopy(financial_data)
    
    # --- Phase 1: Calculate STATIC_ASSUMPTIONS and Inject Value into Cache ---
    print("Phase 1: Calculating Static Assumptions...")
    for metric_name, definition in STATIC_ASSUMPTIONS.items():
        components = definition.get("components")
        operation = definition.get("operation") or definition.get("operator")
        years_to_lookback = definition.get("years", 5)
        
        static_result = np.nan
        
        if operation == "MIN_RATIO_N_YEARS" and len(components) == 2:
            numerator_data = calculated_data_cache.get(components[0], {})
            denominator_data = calculated_data_cache.get(components[1], {})
            
            ratios = []
            
            # Find the ratio for the last N full years
            years_to_check = recent_years_list[:years_to_lookback]
            
            for year in years_to_check:
                num = numerator_data.get(year)
                den = denominator_data.get(year)
                
                # Check for valid, non-zero denominator
                if isinstance(num, (int, float)) and isinstance(den, (int, float)) and den != 0 and not np.isnan(num) and not np.isnan(den):
                    ratios.append(num / den)
                    
            if ratios:
                static_result = min(ratios)
        
        static_results[metric_name] = static_result
        
        # Inject the static result into the calculated_data_cache for use in Phase 2
        # The value is constant across all years for consistency with how other data is structured
        # Use only the years present in the raw data
        all_years = sorted([int(y) for d in financial_data.values() for y in d.keys()], reverse=True)
        calculated_data_cache[metric_name] = {str(year): static_result for year in all_years}

    # --- Phase 2: Calculate TIME-SERIES Metrics ---
    print("Phase 2: Calculating Time-Series Metrics...")
    
    # Iterate through the calculation definitions
    for metric_name, definition in CALCULATED_EQUATIONS.items():
        components = definition.get("components")
        operation = definition.get("operation") or definition.get("operator")
        multiplier = definition.get("multiplier", 1)
        years_to_lookback = definition.get("years")
        result_data = {}

        if not isinstance(components, list) or len(components) == 0:
            continue

        # --- CUSTOM HANDLER: Enterprise Value with optional components ---
        if metric_name == "Enterprise Value":
            # Always require Market Cap; other components are added only if present
            mc_data = calculated_data_cache.get("Market Cap", {})
            if not mc_data:
                time_series_results[metric_name] = {}
                calculated_data_cache[metric_name] = {}
                continue

            # Collect optional component series (may be empty)
            optional_names = ["Net Debt", "Preferred Stock", "Minority Interest"]
            optional_series = {name: calculated_data_cache.get(name, {}) for name in optional_names}

            for year, mc_val in mc_data.items():
                try:
                    total_val = float(mc_val) if mc_val is not None else np.nan
                except Exception:
                    total_val = np.nan

                # Add each optional component if the value exists and is numeric
                for name, series in optional_series.items():
                    if not series:
                        continue
                    comp_val = series.get(year)
                    try:
                        if comp_val is not None and not np.isnan(comp_val):
                            if np.isnan(total_val):
                                total_val = float(comp_val)
                            else:
                                total_val += float(comp_val)
                    except Exception:
                        continue

                result_data[str(year)] = total_val

            time_series_results[metric_name] = result_data
            calculated_data_cache[metric_name] = result_data
            continue

        # --- AVERAGING, CONSOLIDATION, AND CAGR OPERATIONS (Special Case) ---
        if operation == "AVERAGE_PRIOR":
            # Get data from the cache
            data_a = calculated_data_cache.get(components[0], {})
            if not data_a:
                continue
            
            sorted_years_a = sorted(data_a.keys(), key=int)
            
            for i, year_str in enumerate(sorted_years_a):
                current_year = int(year_str)
                if i == 0:
                    result_data[year_str] = np.nan
                    continue
                
                prior_year_str = sorted_years_a[i-1]
                val_a = data_a.get(year_str)
                val_b = data_a.get(prior_year_str)
                
                if val_a is None or val_b is None:
                    result_data[year_str] = np.nan
                    continue
                
                try:
                    val_a = float(val_a)
                    val_b = float(val_b)
                    
                    if val_a is np.nan or val_b is np.nan:
                        result_data[year_str] = np.nan
                        continue
                        
                    result_data[year_str] = ((val_a + val_b) / 2) * multiplier
                except Exception:
                    result_data[year_str] = np.nan
            
        elif operation == "CONSOLIDATE_FILL":
            # Consolidation logic (e.g., Interest Expense)
            data_a = calculated_data_cache.get(components[0], {})
            data_b = calculated_data_cache.get(components[1], {})
            
            # Start with A's data
            result_data = data_a.copy()
            
            # Fill missing keys in A with B's data
            for year, val_b in data_b.items():
                if year not in result_data or np.isnan(result_data.get(year)):
                    result_data[year] = val_b
                    
        elif operation == "CAGR":
            # CAGR logic
            data_a = calculated_data_cache.get(components[0], {})
            if not data_a:
                continue
                
            years_to_lookback = definition.get("years", 5) # Default to 5 years
            
            # Use the most recent year as the end point (Yr N)
            sorted_years = sorted(data_a.keys(), key=int, reverse=True)
            if not sorted_years:
                continue
            
            year_n_str = sorted_years[0] # Most recent year
            
            # Try to find the start year (N-T)
            # T is years_to_lookback, so we need the year T periods ago
            # Example: N=2023, T=5. We need 2023-5 = 2018.
            target_start_year = int(year_n_str) - years_to_lookback
            
            year_0_str = str(target_start_year)
            
            val_n = data_a.get(year_n_str)
            val_0 = data_a.get(year_0_str)

            if val_n is None or val_0 is None or val_0 == 0 or np.isnan(val_n) or np.isnan(val_0):
                cagr = np.nan
            else:
                try:
                    val_n = float(val_n)
                    val_0 = float(val_0)
                    
                    # --- CRITICAL FIX: Check for negative ratio to prevent complex numbers ---
                    ratio = val_n / val_0
                    
                    if ratio < 0:
                        # Cannot calculate real CAGR when the sign changes (e.g., Net Income goes from -100 to 100)
                        # Set to NaN, or 0.0 if you prefer a number, but NaN is safer for a financial model
                        cagr = np.nan
                    else:
                        # CAGR formula: ((Vn / V0)^(1/T) - 1) * Multiplier
                        cagr = (ratio**(1 / years_to_lookback) - 1) * multiplier
                    # --- END FIX ---

                except Exception:
                    cagr = np.nan
            
            # CAGR is a static metric (it doesn't change year-to-year)
            # Store it for the most recent year only for the peer comparison sheet
            result_data[year_n_str] = cagr
            
        # --- SIMPLE MATH OPERATIONS (+, -, *, /) ---
        else:
            # 1. Identify all common years for the components (including constants)
            component_data = {}
            common_years = None
            
            for comp in components:
                try:
                    # If component is a constant (int/float), store it as a value
                    float_val = float(comp)
                    component_data[comp] = float_val
                    continue
                except:
                    pass
                
                # If component is a metric name, get its data from the cache
                data = calculated_data_cache.get(comp, {})
                if not data:
                    result_data = {}
                    break
                    
                component_data[comp] = data
                
                years_set = set(data.keys())
                common_years = years_set if common_years is None else common_years.intersection(years_set)
            
            if not component_data:
                continue
                
            # If no common years were found (e.g., trying to calculate a ratio where one
            # component has no data at all), use all years available in the cache for
            # consistency, where missing values will be handled as NaN.
            if common_years is None:
                all_years = set()
                for key, data in calculated_data_cache.items():
                    if isinstance(data, dict):
                        all_years.update(data.keys())
                common_years = all_years

            sorted_common_years = sorted(list(common_years), key=int)
            
            for year in sorted_common_years:
                try:
                    comp_1 = components[0]
                    val_1 = component_data.get(comp_1)
                    
                    if not isinstance(val_1, dict):
                        calculated_val = float(val_1)
                    else:
                        calculated_val = val_1.get(year)
                    
                    calculated_val = float(calculated_val)
                    
                    if np.isnan(calculated_val):
                        result_data[year] = np.nan
                        continue
                        
                    # Apply subsequent components and operators
                    for i in range(1, len(components)):
                        comp_i = components[i]
                        val_i_raw = component_data.get(comp_i)
                        
                        if not isinstance(val_i_raw, dict):
                            val_i = float(val_i_raw)
                        else:
                            val_i = val_i_raw.get(year)
                            
                        val_i = float(val_i)

                        if val_i is None or np.isnan(val_i):
                            calculated_val = np.nan
                            break

                        operator = operation
                        
                        # --- START MULTI-COMPONENT ADDITION/SUBTRACTION LOGIC ---
                        # Handle NOPAT (Heavy) which is (Op. Income * Tax Ret. Ratio)
                        if metric_name == "Net Operating Profit After Taxes (Heavy)" and operator == '*':
                            if i == 1:
                                calculated_val *= val_i
                                break # Done with this calculation
                        # --- END MULTI-COMPONENT ADDITION/SUBTRACTION LOGIC ---

                        elif operator == '+':
                            calculated_val += val_i
                        elif operator == '-':
                            calculated_val -= val_i
                        elif operator == '/':
                            calculated_val = (calculated_val / val_i) if val_i != 0 else np.nan
                        elif operator == '*':
                            calculated_val *= val_i
                            
                    # Apply final multiplier (e.g., *100 for percentage)
                    if not np.isnan(calculated_val):
                        result_data[year] = calculated_val * multiplier
                    else:
                        result_data[year] = np.nan
                        
                except Exception:
                    result_data[year] = np.nan

        # 2. Store the result and inject into the cache for use by subsequent formulas
        time_series_results[metric_name] = result_data
        calculated_data_cache[metric_name] = result_data

    # --- Ensure direct raw metrics required for export are present ---
    for metric_name in CALCULATED_EXPORT_ORDER:
        if metric_name not in time_series_results and metric_name in financial_data:
            time_series_results[metric_name] = financial_data[metric_name]
            calculated_data_cache[metric_name] = financial_data[metric_name]

    # --- Finalize Time Series DataFrame ---
    calculated_metrics_df = pd.DataFrame(time_series_results).transpose()
    if not calculated_metrics_df.empty:
        # Filter to only keep year columns
        calculated_metrics_df = calculated_metrics_df[[col for col in calculated_metrics_df.columns if str(col).isdigit()]]
        # Reindex to match the desired export order
        calculated_metrics_df = calculated_metrics_df.reindex(CALCULATED_EXPORT_ORDER)
        # Sort years descending (most recent first)
        calculated_metrics_df = calculated_metrics_df.sort_index(axis=1, ascending=False)

    return calculated_metrics_df, static_results

# --- New Function: Calculate Peer Group Price Metrics (Unchanged) ---
def calculate_price_metrics_and_beta(tickers, shares_outstanding_map):
    """
    Fetches daily price data from yfinance, calculates 90-day SMA, Market Cap, and Beta
    relative to the peer group (equally-weighted average).
    Returns two consolidated DataFrames: price_history_df and summary_df.
    """
    print("\n--- Calculating Peer Group Price Metrics and Beta ---")
    
    # 1. Fetch 1 year of daily price data for all tickers
    end_date = date.today().strftime('%Y-%m-%d')
    # Use 370 days to ensure at least 1 year of trading days
    start_date = (date.today() - relativedelta(days=370)).strftime('%Y-%m-%d')
    
    # Fetch all data at once
    try:
        data = yf.download(tickers, start=start_date, end=end_date, progress=False)
        if data.empty:
            print("Error: Could not fetch daily price data from Yahoo Finance.")
            return pd.DataFrame(), pd.DataFrame()
            
        close_prices = data['Close']
    except Exception as e:
        print(f"Error fetching daily price data: {e}")
        return pd.DataFrame(), pd.DataFrame()

    # Remove tickers that may have failed to fetch (NaN columns)
    close_prices = close_prices.dropna(axis=1, how='all')
    tickers = close_prices.columns.tolist()
    
    if not tickers:
        print("Error: No valid ticker price data found for calculation.")
        return pd.DataFrame(), pd.DataFrame()

    # 2. Calculate Daily Returns for all stocks
    daily_returns = close_prices.pct_change().dropna()

    # 3. Calculate Benchmark Daily Return (Equally-weighted average of all peers)
    # The benchmark is the average of all the daily returns in the current list of tickers
    benchmark_returns = daily_returns.mean(axis=1)

    # 4. Prepare results containers
    summary_data = {}
    price_history_frames = []

    # 5. Iterate through each stock for individual metrics
    for ticker in tickers:
        stock_returns = daily_returns[ticker].dropna()
        stock_prices = close_prices[ticker].dropna()
        
        # Ensure we have enough data to proceed
        if stock_returns.empty or ticker not in close_prices.columns:
            print(f"Skipping {ticker}: Insufficient daily return data.")
            continue

        # --- A. Calculate Beta ---
        beta = np.nan
        # Filter benchmark returns to match the index of the stock's returns
        common_index = stock_returns.index.intersection(benchmark_returns.index)
        
        if len(common_index) > 1:
            stock_returns_aligned = stock_returns.loc[common_index]
            benchmark_returns_aligned = benchmark_returns.loc[common_index]
            
            # Beta formula: Covariance(Stock Returns, Benchmark Returns) / Variance(Benchmark Returns)
            if benchmark_returns_aligned.var() != 0:
                beta = stock_returns_aligned.cov(benchmark_returns_aligned) / benchmark_returns_aligned.var()
            
        # --- B. Calculate 90-Day Simple Moving Average (SMA) ---
        sma_90 = stock_prices.rolling(window=90).mean()
        
        # Get the most recent 90-day average price
        avg_price_90_day = sma_90.iloc[-1] if not sma_90.empty else np.nan
        
        # --- C. Calculate Market Cap (90-Day Avg.) ---
        most_recent_shares = shares_outstanding_map.get(ticker, 0.0)
        market_cap_90_day = np.nan
        
        # Ensure scalar values for comparison
        if hasattr(avg_price_90_day, "iloc"):
            avg_price_90_day = avg_price_90_day.iloc[-1]
        if hasattr(most_recent_shares, "iloc"):
            most_recent_shares = most_recent_shares.iloc[-1]

        # Validate before calculating
        if not np.isnan(avg_price_90_day) and most_recent_shares > 0:
            # Market Cap (90-Day Avg.) = 90-Day Average Price * Total Shares Outstanding
            market_cap_90_day = avg_price_90_day * most_recent_shares

        # --- D. Store Summary Metrics ---
        summary_data[ticker] = {
            'Beta': beta,
            '90-Day Average Price': avg_price_90_day,
            'Total Common Shares Outstanding (M)': most_recent_shares / 1e6 if not np.isnan(most_recent_shares) else np.nan, # Show in Millions
            'Market Cap (90-Day Avg.) (B)': market_cap_90_day / 1e9 if not np.isnan(market_cap_90_day) else np.nan # Show in Billions
        }

        # --- E. Prepare Price History DataFrames for Consolidation ---
        # Select the columns for the current stock
        price_df = pd.DataFrame({
            f'{ticker} - Close Price': stock_prices,
            f'{ticker} - 90-Day SMA': sma_90
        })
        price_history_frames.append(price_df)

    # 6. Create Final DataFrames
    # Consolidated Price History (Outer join on index to keep all dates)
    price_history_df = pd.concat(price_history_frames, axis=1).sort_index(ascending=False)
    price_history_df.index.name = "Date"

    # Summary DataFrame
    summary_df = pd.DataFrame(summary_data).transpose()
    summary_df.index.name = "Ticker"

    # Re-order the columns for the summary sheet
    summary_export_order = [
        'Beta',
        '90-Day Average Price',
        'Total Common Shares Outstanding (M)',
        'Market Cap (90-Day Avg.) (B)'
    ]
    summary_df = summary_df.reindex(columns=summary_export_order)
    
    print("Price metrics and Beta calculated successfully.")
    return price_history_df, summary_df

# --- NEW FUNCTION: DataFrame Cleaning for Export (FIXED COMPLEX NUMBER ISSUE) ---
def clean_dataframe_for_export(df):
    """
    Cleans the DataFrame by ensuring columns are strings, replacing NaN/None with np.nan,
    and converting complex numbers (like 0j) to their real part (0.0).
    """
    df.columns = df.columns.astype(str)
    
    # 1. Replace None with NaN
    df = df.replace({None: np.nan})
    
    # 2. Iterate and apply casting/rounding to the data to remove complex numbers and clean up floats
    for col in df.columns:
        
        # Check if the series contains complex numbers and convert them to their real part
        if np.iscomplexobj(df[col].dtypes):
            # This converts complex(x, y) to float(x). If y is 0, it removes the 'j'.
            df[col] = df[col].apply(lambda x: x.real if isinstance(x, complex) else x)
        
        # 3. Rounding to a high precision (6 decimal places) for storage consistency
        df[col] = df[col].apply(lambda x: round(x, 6) if isinstance(x, (float, np.float64)) and not np.isnan(x) else x)

    return df

# --- NEW FUNCTION: Create and Save the ALL FINANCIALS Book (Unchanged) ---
def create_and_save_all_financials_book(consolidated_raw_dfs, consolidated_calcs, static_results_map, file_path):
    """ Writes all raw and calculated data for ALL tickers to a single Excel workbook. """
    print("\n--- STEP A: Creating All Financial Statements Book ---")
    
    if not consolidated_raw_dfs or not consolidated_calcs:
        print("Skipping All Financial Statements Book: No data collected.")
        return

    try:
        # Use ExcelWriter with openpyxl engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for ticker in sorted(consolidated_raw_dfs.keys()):
                raw_df = consolidated_raw_dfs.get(ticker)
                calc_df = consolidated_calcs.get(ticker)
                
                # Write RAW data sheet
                if raw_df is not None and not raw_df.empty:
                    raw_df = clean_dataframe_for_export(raw_df)
                    raw_df.to_excel(writer, sheet_name=f"{ticker}_RAW", index=True)
                
                # Write CALCULATED data sheet
                if calc_df is not None and not calc_df.empty:
                    calc_df = clean_dataframe_for_export(calc_df)
                    calc_df.to_excel(writer, sheet_name=f"{ticker}_CALCULATED", index=True)
        
        # Now, re-open the file to add static assumptions to the calculated sheets using openpyxl
        if os.path.exists(file_path) and static_results_map:
            book = load_workbook(file_path)
            for ticker, static_results in static_results_map.items():
                sheet_name_calculated = f"{ticker}_CALCULATED"
                if sheet_name_calculated in book.sheetnames and static_results:
                    ws = book[sheet_name_calculated]
                    
                    # Find the first empty row below the data
                    max_row = ws.max_row + 2
                    
                    ws[f'A{max_row}'] = "STATIC ASSUMPTIONS (Used for Light ROIC Modeling)"
                    max_row += 1
                    
                    for name, value in static_results.items():
                        ws[f'A{max_row}'] = name
                        ws[f'B{max_row}'] = value
                        max_row += 1
                        
            book.save(file_path)

        print(f"All Financial Statements Book saved: {file_path}")
        
        # Automatic opening of the file
        if os.path.exists(file_path):
            print(f"Attempting to automatically open the Excel file: {file_path}")
            try:
                os.system(f'open "{file_path}"')
            except Exception as e:
                print(f"Warning: Could not automatically open {file_path}. Error: {e}")

    except Exception as e:
        print(f"\nCRITICAL CONSOLIDATED FINANCIALS SAVE ERROR: An error occurred while writing the All Financials book: {e}")

# --- LIST OF CAGR Keys (Unchanged) ---
REQUIRED_CAGR_KEYS = [
    "Interest Bearing Debt CAGR (3-Year) (%)",
    "ROIC (Light) CAGR (3-Year) (%)",
    "ROIC (Heavy) CAGR (3-Year) (%)",
    "Net Profit Margin (%) CAGR (3-Year) (%)",
    "Asset Turnover Ratio (%) CAGR (3-Year)",
    "Tax Retention Ratio (%) CAGR (3-Year)",
    "Interest Coverage Ratio (%) CAGR (3-Year)",
    "Quick Ratio 'Liquid Ratio' (%) CAGR (3-Year)",
    "Adjusted Debt to Asset Ratio (%) CAGR (3-Year)",
    "Adjusted Debt to Equity Ratio (%) CAGR (3-Year)",
    "Shareholder/Stockholder Equity (%) CAGR (3-Year)",
    "Cash From Operations (%) CAGR (3-Year)",
    "Earnings Per Share (EPS) CAGR (3-Year) (%)",
    "Gross Profit Margin (%) CAGR (3-Year) (%)",
    "Operating Profit Margin (%) CAGR (3-Year) (%)",
    "Return On Assets (%) CAGR (3-Year) (%)",
    "Return On Equity (%) CAGR (3-Year)",
    "Value Spread (Heavy) (%) CAGR (3-Year)",
    "Value Spread (Light) (%) CAGR (3-Year)",
    "Market Cap (%) CAGR (3-Year)",
    "Enterprise Value (%) CAGR (3-Year)",
    "Tangible Book Value Per Share (%) CAGR (3-Year)",
    "P/E Ratio (Calculated) (%) CAGR (3-Year)",
    "Price To Book (P/B Ratio) (%) CAGR (3-Year)",
    "EV / EBITDA (%) CAGR (3-Year)",
    "EV / Share (%) CAGR (3-Year)",
    "Total Revenue (%) CAGR (3-Year)",
    "Net Income (%) CAGR (3-Year)",
    "Total Assets (%) CAGR (3-Year)",
    
]

# --- NEW FUNCTION: Create and Save the PEER COMPARISON Book (Unchanged) ---

# --- Metric groups for Peer Summary Statistics sheet ---
# --- Configuration for Peer Summary Statistics sheet ---
# Edit this list to change which metrics appear, their grouping, and
# whether they should have a 3-year CAGR computed in the RIGHT table.
#
# - "group": the section title row (shown in quotes in your screenshot).
# - {"group": ""} creates a blank spacer row between groups.
# - "name": EXACT metric label as it appears in Comparison Peer Relevant Data.
# - "cagr": True  -> metric appears in the 3Y CAGR table (right side).
#          False -> metric will have blank CAGR stats.

PEER_SUMMARY_CONFIG = [
    {
        "group": "Profitability Ratios Metrics",
        "metrics": [
            {"name": "Gross Profit Margin (%)",       "cagr": True},
            {"name": "Operating Profit Margin (%)",   "cagr": True},
            {"name": "Net Profit Margin (%)",         "cagr": True},
            {"name": "Return On Assets",              "cagr": True},
            {"name": "Return On Equity",              "cagr": True},
            {"name": "ROIC (Light)",                  "cagr": True},
            {"name": "ROIC (Heavy)",                  "cagr": True},
            {"name": "WACC (%)",                      "cagr": True},
            {"name": "Value Spread (Heavy) (%)",      "cagr": True},
            {"name": "Value Spread (Light) (%)",      "cagr": True},
        ],
    },
    {"group": ""},  # spacer row

    {
        "group": "Price Valuation Metrics",
        "metrics": [
            {"name": "Closing Price (USD)",           "cagr": True},
            {"name": "Market Cap",                    "cagr": True},
            {"name": "Enterprise Value",              "cagr": True},
            {"name": "Earnings Per Share (EPS)",      "cagr": True},
            {"name": "Tangible Book Value Per Share", "cagr": True},
            {"name": "P/E Ratio (Calculated)",        "cagr": True},
            {"name": "Price To Book (P/B Ratio)",     "cagr": True},
            {"name": "EV / EBITDA",                   "cagr": True},
            {"name": "EV / Share",                    "cagr": True},
            {"name": "PEG Ratio (Calculated)",        "cagr": True},
        ],
    },
    {"group": ""},  # spacer row

    {
        "group": "Base SEC Metrics",
        "metrics": [
            {"name": "Total Revenue",                 "cagr": True},
            {"name": "Net Income",                    "cagr": True},
            {"name": "Total Assets",                  "cagr": True},
            {"name": "Cash From Operations",          "cagr": True},
            {"name": "Free Cash Flow",                "cagr": True},
            {"name": "Interest Bearing Debt",         "cagr": True},
            {"name": "Shareholder/Stockholder Equity","cagr": True},
        ],
    },
    {"group": ""},  # spacer row

    {
        "group": "Solvency Ratios",
        "metrics": [
            {"name": "Adjusted Debt to Equity Ratio", "cagr": True},
            {"name": "Adjusted Debt to Asset Ratio",  "cagr": True},
            {"name": "Quick Ratio 'Liquid Ratio'",    "cagr": True},
            {"name": "Interest Coverage Ratio",       "cagr": True},
        ],
    },
    {"group": ""},  # spacer row
    
    {
        "group": "Fair Value Screen Data",
        "metrics": [
            {"name": "EV / FCF Forward (3Y)", "cagr": True},
            {"name": "Gross Profit / EV", "cagr": True},
            {"name": "Price / Sales","cagr": True},
        ],
    },
    {"group": ""},
    
    {
        "group": "Other",
        "metrics": [
            {"name": "Tax Retention Ratio",           "cagr": True},
            {"name": "Asset Turnover Ratio",          "cagr": True},
        ],
    },
]

def create_and_save_comparison_book(consolidated_calcs, price_history_df, summary_df, file_path):
    """
    Writes consolidated comparison metrics (ratios), CAGR, price history, and
    price summary data to a single Excel workbook.
    """
    print("\n--- STEP B: Creating Peer Comparison Book ---")

    # 1. Comparison Metrics Sheet (Ratios)
    comparison_dfs = []
    # Create the comparison DF by selecting the most recent year's data (first column)
    for ticker, df in consolidated_calcs.items():
        if df.columns.empty:
            continue
        # Ensure year columns are strings for safe indexing
        df.columns = df.columns.astype(str)
        most_recent_year = df.columns[0]
        # Select the column corresponding to the most recent year
        comparison_dfs.append(df[most_recent_year].rename(ticker))
        
    # Metrics as rows, tickers as columns (Option B)
    comparison_df = pd.concat(comparison_dfs, axis=1)
    comparison_df.index.name = "Metric"

    # 2. CAGR Sheet (Simplified to extract pre-calculated metrics)
    cagr_data = {}
    for ticker, df in consolidated_calcs.items():
        ticker_cagrs = {}
        # Ensure year columns are strings for safe indexing
        df.columns = df.columns.astype(str)
        
        # Handle case where DataFrame might be empty after re-indexing
        if df.columns.empty:
            continue
            
        most_recent_year = df.columns[0]
        
        # Iterate over the new list of REQUIRED_CAGR_KEYS (pre-calculated metrics)
        for cagr_key in REQUIRED_CAGR_KEYS:
            if cagr_key in df.index:
                # Extract the value for the most recent year
                ticker_cagrs[cagr_key] = df.loc[cagr_key, most_recent_year]
                
        cagr_data[ticker] = pd.Series(ticker_cagrs)
        
    # Metrics as rows, tickers as columns (Option B)
    cagr_df = pd.DataFrame(cagr_data)
    cagr_df.index.name = "Metric"

    # 2.5 Inject dynamic 90-day Market Cap and Enterprise Value into comparison_df
    # We prefer price-derived Market Cap (90-Day Avg.) over any static SEC-based value.
    for ticker in comparison_df.columns:
        # Default placeholders
        mcap_90_raw = np.nan
        net_debt_latest = np.nan
        shares_latest = np.nan

        # A) Get 90-day Market Cap from summary_df (stored in Billions)
        try:
            if ticker in summary_df.index and "Market Cap (90-Day Avg.) (B)" in summary_df.columns:
                mcap_B = summary_df.loc[ticker, "Market Cap (90-Day Avg.) (B)"]
                if mcap_B is not None and not (isinstance(mcap_B, float) and np.isnan(mcap_B)):
                    mcap_90_raw = float(mcap_B) * 1e9
        except Exception:
            pass

        # B) Get latest Net Debt and Shares Outstanding from the calculated sheets
        try:
            df_calc = consolidated_calcs.get(ticker)
            if df_calc is not None and not df_calc.empty:
                df_calc.columns = df_calc.columns.astype(str)

                # Net Debt (latest non-NaN year)
                if "Net Debt" in df_calc.index:
                    nd_row = df_calc.loc["Net Debt", :]
                    nd_non_na = nd_row.dropna()
                    if not nd_non_na.empty:
                        net_debt_latest = float(nd_non_na.iloc[0])

                # Shares outstanding (latest non-NaN year)
                if "Total Common Shares Outstanding" in df_calc.index:
                    sh_row = df_calc.loc["Total Common Shares Outstanding", :]
                    sh_non_na = sh_row.dropna()
                    if not sh_non_na.empty:
                        shares_latest = float(sh_non_na.iloc[0])
        except Exception:
            pass

        # If shares still NaN, fall back to summary_df "Total Common Shares Outstanding (M)"
        if (not isinstance(shares_latest, (int, float)) or np.isnan(shares_latest)) and \
           ticker in summary_df.index and "Total Common Shares Outstanding (M)" in summary_df.columns:
            try:
                sh_M = summary_df.loc[ticker, "Total Common Shares Outstanding (M)"]
                if sh_M is not None and not (isinstance(sh_M, float) and np.isnan(sh_M)):
                    shares_latest = float(sh_M) * 1e6
            except Exception:
                pass

        # C) Compute dynamic Enterprise Value from 90-day Market Cap and latest Net Debt
        ev_dynamic = np.nan
        if isinstance(mcap_90_raw, (int, float)) and not np.isnan(mcap_90_raw) and \
           isinstance(net_debt_latest, (int, float)) and not np.isnan(net_debt_latest):
            ev_dynamic = mcap_90_raw + net_debt_latest

        # D) Push these values into comparison_df where appropriate
        if "Market Cap" in comparison_df.index and isinstance(mcap_90_raw, (int, float)) and not np.isnan(mcap_90_raw):
            comparison_df.at["Market Cap", ticker] = mcap_90_raw

        if "Enterprise Value" in comparison_df.index and isinstance(ev_dynamic, (int, float)) and not np.isnan(ev_dynamic):
            comparison_df.at["Enterprise Value", ticker] = ev_dynamic

        if "EV / Share" in comparison_df.index and \
           isinstance(ev_dynamic, (int, float)) and not np.isnan(ev_dynamic) and \
           isinstance(shares_latest, (int, float)) and not np.isnan(shares_latest) and shares_latest > 0:
            comparison_df.at["EV / Share", ticker] = ev_dynamic / shares_latest

        if "EV / EBITDA" in comparison_df.index and isinstance(ev_dynamic, (int, float)) and not np.isnan(ev_dynamic):
            try:
                if "EBITDA" in comparison_df.index:
                    ebitda_val = comparison_df.at["EBITDA", ticker]
                    ebitda_val = float(ebitda_val)
                    if not np.isnan(ebitda_val) and ebitda_val != 0:
                        comparison_df.at["EV / EBITDA", ticker] = ev_dynamic / ebitda_val
            except Exception:
                pass

        # === NEW: WACC and Value Spread Calculations ===
        try:
            # 1) Cost of Equity (Re) via CAPM
            cost_of_equity = np.nan
            if ticker in summary_df.index and "Beta" in summary_df.columns:
                beta_val = summary_df.loc[ticker, "Beta"]
                if isinstance(beta_val, (int, float)) and not np.isnan(beta_val):
                    cost_of_equity = RISK_FREE_RATE + beta_val * EQUITY_RISK_PREMIUM  # decimal

            # 2) Cost of Debt (Rd) ≈ Interest Expense Consolidated / Net Debt, floored at MIN_COST_OF_DEBT
            cost_of_debt = np.nan
            df_calc_for_wacc = consolidated_calcs.get(ticker)
            if df_calc_for_wacc is not None and not df_calc_for_wacc.empty:
                df_calc_for_wacc.columns = df_calc_for_wacc.columns.astype(str)
                if "Interest Expense Consolidated" in df_calc_for_wacc.index:
                    ie_row = df_calc_for_wacc.loc["Interest Expense Consolidated", :].dropna()
                    if not ie_row.empty and isinstance(net_debt_latest, (int, float)) and net_debt_latest > 0 and not np.isnan(net_debt_latest):
                        latest_ie = float(ie_row.iloc[0])
                        cost_of_debt = latest_ie / net_debt_latest

            # Apply floor to Rd, and ensure it's numeric
            if not isinstance(cost_of_debt, (int, float)) or np.isnan(cost_of_debt) or cost_of_debt < MIN_COST_OF_DEBT:
                cost_of_debt = MIN_COST_OF_DEBT

            # 3) Capital structure weights using dynamic Market Cap and Net Debt
            equity_weight = np.nan
            debt_weight = np.nan
            if isinstance(mcap_90_raw, (int, float)) and not np.isnan(mcap_90_raw) and \
               isinstance(net_debt_latest, (int, float)) and not np.isnan(net_debt_latest):
                total_capital = mcap_90_raw + net_debt_latest
                if total_capital > 0:
                    equity_weight = mcap_90_raw / total_capital
                    debt_weight = net_debt_latest / total_capital

            # 4) Tax rate from calculated metrics
            tax_rate = np.nan
            if df_calc_for_wacc is not None and "Tax Rate" in df_calc_for_wacc.index:
                tr_row = df_calc_for_wacc.loc["Tax Rate", :].dropna()
                if not tr_row.empty:
                    tr_val = tr_row.iloc[0]
                    if isinstance(tr_val, (int, float)):
                        tax_rate = float(tr_val)

            # 5) WACC (decimal then convert to %)
            wacc_pct = np.nan
            if all(isinstance(x, (int, float)) and not np.isnan(x) for x in
                   [cost_of_equity, cost_of_debt, equity_weight, debt_weight, tax_rate]):
                wacc_decimal = equity_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - tax_rate)
                wacc_pct = wacc_decimal * 100.0

            # Store WACC components
            comparison_df.at["Cost of Equity (Re)", ticker] = cost_of_equity * 100.0 if isinstance(cost_of_equity, (int, float)) and not np.isnan(cost_of_equity) else np.nan
            comparison_df.at["Cost of Debt (Rd)", ticker] = cost_of_debt * 100.0 if isinstance(cost_of_debt, (int, float)) and not np.isnan(cost_of_debt) else np.nan
            comparison_df.at["Equity Weight", ticker] = equity_weight
            comparison_df.at["Debt Weight", ticker] = debt_weight
            comparison_df.at["WACC (%)", ticker] = wacc_pct

            # 6) Value Spread (Heavy / Light) in percentage points
            vs_heavy = np.nan
            vs_light = np.nan

            roic_heavy_raw = comparison_df.at["ROIC (Heavy)", ticker] if "ROIC (Heavy)" in comparison_df.index else np.nan
            roic_light_raw = comparison_df.at["ROIC (Light)", ticker] if "ROIC (Light)" in comparison_df.index else np.nan

            # ROICs in your pipeline are decimals (e.g., 0.12), convert to %
            roic_heavy_pct = roic_heavy_raw * 100.0 if isinstance(roic_heavy_raw, (int, float)) and not np.isnan(roic_heavy_raw) else np.nan
            roic_light_pct = roic_light_raw * 100.0 if isinstance(roic_light_raw, (int, float)) and not np.isnan(roic_light_raw) else np.nan

            if not np.isnan(roic_heavy_pct) and not np.isnan(wacc_pct):
                vs_heavy = roic_heavy_pct - wacc_pct
            if not np.isnan(roic_light_pct) and not np.isnan(wacc_pct):
                vs_light = roic_light_pct - wacc_pct

            comparison_df.at["Value Spread (Heavy) (%)", ticker] = vs_heavy
            comparison_df.at["Value Spread (Heavy) (%)", ticker] = vs_heavy
            comparison_df.at["Value Spread (Light) (%)", ticker] = vs_light

            # --- NEW: also write WACC and Value Spreads into each ticker's CALCULATED sheet for history/CAGR ---
            try:
                if df_calc_for_wacc is not None and not df_calc_for_wacc.empty and not np.isnan(wacc_pct):
                    # ensure column labels are strings so we can detect year columns
                    df_calc_for_wacc.columns = df_calc_for_wacc.columns.astype(str)
                    year_cols = [c for c in df_calc_for_wacc.columns if str(c).isdigit()]
                    if year_cols:
                        # constant WACC across all historical years (standard modeling assumption)
                        for y in year_cols:
                            df_calc_for_wacc.at["WACC (%)", y] = wacc_pct
                        # value spreads per year = ROIC (Heavy/Light) % - WACC %
                        if "ROIC (Heavy)" in df_calc_for_wacc.index:
                            roic_heavy_series = df_calc_for_wacc.loc["ROIC (Heavy)", year_cols].astype(float) * 100.0
                            df_calc_for_wacc.loc["Value Spread (Heavy) (%)", year_cols] = roic_heavy_series - wacc_pct
                        if "ROIC (Light)" in df_calc_for_wacc.index:
                            roic_light_series = df_calc_for_wacc.loc["ROIC (Light)", year_cols].astype(float) * 100.0
                            df_calc_for_wacc.loc["Value Spread (Light) (%)", year_cols] = roic_light_series - wacc_pct
                        # push updated frame back into consolidated_calcs so it is exported later
                        consolidated_calcs[ticker] = df_calc_for_wacc
            except Exception as _wacc_hist_err:
                print(f"Warning: could not backfill WACC history for {ticker}: {_wacc_hist_err}")
        except Exception as e:
            print(f"WACC/Value Spread computation error for {ticker}: {e}")

    # --- 3. Apply Cleaning and Save to Excel ---
    comparison_df = clean_dataframe_for_export(comparison_df)
    cagr_df = clean_dataframe_for_export(cagr_df)
    price_history_df = clean_dataframe_for_export(price_history_df)
    summary_df = clean_dataframe_for_export(summary_df)

        # --- Build Peer Summary Statistics sheet (latest metrics + 3-year CAGR) ---
    peer_summary_rows = []

    # Helper: convert any 1D-like structure into a clean numeric vector
    def _to_numeric_vector(values):
        """Convert any 1D/2D input (Series, list, ndarray, DataFrame) into
        a clean 1D numeric vector with NaNs and non-numerics removed.
        If a 2D structure is passed, the first row is used.
        """
        vals = values
        # If this is a DataFrame, use the first row (e.g., most recent year across tickers)
        if isinstance(vals, pd.DataFrame):
            if vals.shape[0] > 1:
                vals = vals.iloc[0]
            else:
                vals = vals.iloc[0]
        # If it's a 2D ndarray, take the first row
        elif isinstance(vals, np.ndarray) and vals.ndim > 1:
            vals = vals[0, :]
        # Flatten into a Series and coerce to numeric
        vals = pd.Series(vals).dropna()
        vals = pd.to_numeric(vals, errors="coerce").dropna()
        return vals


    # Helper: compute per-ticker 3-year CAGR values for a given metric name
    
    # Helper: compute per-ticker 3-year CAGR values for a given metric name
    def _compute_metric_cagrs(metric_name, years=3):
        """
        Compute per‑ticker CAGR for a given metric across the consolidated_calcs
        DataFrames, using a sliding window that:
          - Always uses a fixed `years` span (e.g. 3 years => 4 data points).
          - Skips blanks / zeros at the start or end of the window by shifting
            both ends together toward older years.
          - Uses absolute values for the geometric CAGR math, but restores the
            sign based on whether the ending value is better or worse than the
            starting value (end < start => negative CAGR, else positive).
        """
        cagr_values = []

        for ticker, df_calc in consolidated_calcs.items():
            if df_calc is None or df_calc.empty:
                continue
            if metric_name not in df_calc.index:
                continue

            row = df_calc.loc[metric_name]

            # Collect (year, value) pairs for columns that look like years and are numeric.
            numeric_years = []
            for col_label, val in row.items():
                # Column label must be interpretable as an integer year.
                try:
                    year_int = int(str(col_label))
                except Exception:
                    continue
                # Value must be numeric; treat non‑numeric as missing.
                try:
                    v = float(val)
                except Exception:
                    continue
                numeric_years.append((year_int, v))

            # Need at least years+1 data points (e.g., 4 points for a 3‑year CAGR).
            if len(numeric_years) < years + 1:
                continue

            # Sort by year descending so index 0 is the most recent year.
            numeric_years.sort(key=lambda x: x[0], reverse=True)

            cagr_for_ticker = None

            # Sliding window: (offset, offset+years) keeps a fixed `years` span.
            # This implements the "shift both ends together" rule:
            #  2025→2022, if invalid -> 2024→2021, etc.
            max_offset = len(numeric_years) - (years + 1)
            for offset in range(max_offset + 1):
                end_year, end_val = numeric_years[offset]
                start_year, start_val = numeric_years[offset + years]

                # Skip if either end is zero (user wants zero/blank treated as invalid).
                try:
                    end_val_f = float(end_val)
                    start_val_f = float(start_val)
                except Exception:
                    continue

                if end_val_f == 0 or start_val_f == 0:
                    # Try the next older window.
                    continue

                # Use absolute values for the geometric CAGR math.
                abs_start = abs(start_val_f)
                abs_end = abs(end_val_f)

                if abs_start == 0 or abs_end == 0:
                    continue

                try:
                    raw_cagr = (abs_end / abs_start) ** (1.0 / years) - 1.0
                except Exception:
                    continue

                # Directional sign rule:
                #   - If end < start  -> negative CAGR (performance worsened)
                #   - Else           -> positive CAGR (performance improved or flat)
                if end_val_f < start_val_f:
                    cagr_for_ticker = -abs(raw_cagr)
                else:
                    cagr_for_ticker = abs(raw_cagr)

                # We found the most recent valid window; stop searching older ones.
                break

            # Only record if we successfully computed a CAGR for this ticker.
            if cagr_for_ticker is not None:
                cagr_values.append(cagr_for_ticker)

        return pd.Series(cagr_values)

    for group_cfg in PEER_SUMMARY_CONFIG:
        group_name = group_cfg.get("group", "")
        metrics = group_cfg.get("metrics", []) or []

        # Spacer row: blank group with no metrics
        if not group_name and not metrics:
            peer_summary_rows.append({
                "Metric": "",
                "Relevant Mean of ALL Stocks": np.nan,
                "Standard Deviation": np.nan,
                "COV": np.nan,
                "": "",
                "Metric (3Y CAGR)": "",
                "3Y CAGR Mean of ALL Stocks": np.nan,
                "3Y CAGR Standard Deviation": np.nan,
                "3Y CAGR COV": np.nan,
            })
            continue

        # Group title row
        if group_name:
            peer_summary_rows.append({
                "Metric": group_name,
                "Relevant Mean of ALL Stocks": np.nan,
                "Standard Deviation": np.nan,
                "COV": np.nan,
                "": "",
                "Metric (3Y CAGR)": group_name,
                "3Y CAGR Mean of ALL Stocks": np.nan,
                "3Y CAGR Standard Deviation": np.nan,
                "3Y CAGR COV": np.nan,
            })

        # Actual metric rows
        for metric_cfg in metrics:
            metric_name = metric_cfg.get("name")
            use_cagr = bool(metric_cfg.get("cagr", True))

            # LEFT TABLE: most recent values from comparison_df
            if metric_name in comparison_df.index:
                series_left = comparison_df.loc[metric_name]
                valid_left = _to_numeric_vector(series_left)
                if not valid_left.empty:
                    mean_left = float(valid_left.mean())
                    std_left = float(valid_left.std(ddof=0))
                    cv_left = (std_left / mean_left) if mean_left != 0 else np.nan
                else:
                    mean_left = std_left = cv_left = np.nan
            else:
                mean_left = std_left = cv_left = np.nan

            # RIGHT TABLE: 3-year CAGR, per metric, across all tickers
            if use_cagr:
                valid_right = _compute_metric_cagrs(metric_name, years=3)
                if not valid_right.empty:
                    mean_right = float(valid_right.mean())
                    std_right = float(valid_right.std(ddof=0))
                    cv_right = (std_right / mean_right) if mean_right != 0 else np.nan
                else:
                    mean_right = std_right = cv_right = np.nan
                # Use the base metric name as the label so downstream sheets can match on it
                cagr_label = metric_name
            else:
                mean_right = std_right = cv_right = np.nan
                cagr_label = ""

            peer_summary_rows.append({
                "Metric": metric_name,
                "Relevant Mean of ALL Stocks": mean_left,
                "Standard Deviation": std_left,
                "COV": cv_left,
                "": "",
                "Metric (3Y CAGR)": cagr_label,
                "3Y CAGR Mean of ALL Stocks": mean_right,
                "3Y CAGR Standard Deviation": std_right,
                "3Y CAGR COV": cv_right,
            })

    peer_summary_df = pd.DataFrame(peer_summary_rows)
        # === Helper: scoring for valuation metrics (Lower vs Higher is better) ===
    def _score_lower_is_better(ticker_val, bench_val):
        """
        For valuation ratios where a LOWER value is better (e.g. EV/EBITDA, P/E, PEG).
        Returns positive score when ticker < benchmark, negative when ticker > benchmark.
        """
        if not isinstance(ticker_val, (int, float)) or not isinstance(bench_val, (int, float)):
            return np.nan
        if np.isnan(ticker_val) or np.isnan(bench_val) or bench_val == 0:
            return np.nan
        return (bench_val - ticker_val) / abs(bench_val) * 100.0

    def _score_higher_is_better(ticker_val, bench_val):
        """
        For metrics where a HIGHER value is better (e.g. Value Spread).
        Returns positive score when ticker > benchmark, negative when ticker < benchmark.
        """
        if not isinstance(ticker_val, (int, float)) or not isinstance(bench_val, (int, float)):
            return np.nan
        if np.isnan(ticker_val) or np.isnan(bench_val) or bench_val == 0:
            return np.nan
        return (ticker_val - bench_val) / abs(bench_val) * 100.0

        # Helper: map base metric names (e.g. "Net Profit Margin (%)")
    # to the corresponding CAGR row name in Peer Growth Summary (cagr_df),
    # e.g. "Net Profit Margin (%) CAGR (3-Year) (%)".
    metric_to_cagr_row = {}

    def _get_cagr_row_name(metric_name):
        """
        Find the CAGR row name in cagr_df that corresponds to a base metric name.
        We match rows where:
          - The row index starts with the base metric name, and
          - The text 'CAGR' appears in the row label.
        Returns the matching row label or None.
        """
        if metric_name in metric_to_cagr_row:
            return metric_to_cagr_row[metric_name]

        row_name = None
        for idx in cagr_df.index:
            try:
                idx_str = str(idx)
                if idx_str.startswith(metric_name) and "CAGR" in idx_str:
                    row_name = idx_str
                    break
            except Exception:
                continue

        metric_to_cagr_row[metric_name] = row_name
        return row_name

    def _get_peer_growth_cagr(metric_name, ticker):
        """
        Return the ticker's CAGR value for the given base metric name,
        using the Peer Growth Summary (cagr_df) as the single source of truth.
        """
        row_name = _get_cagr_row_name(metric_name)
        if row_name is None:
            return np.nan
        if ticker not in cagr_df.columns:
            return np.nan
        try:
            val = cagr_df.at[row_name, ticker]
            return float(val)
        except Exception:
            return np.nan



    try:
        # Using openpyxl to ensure robust handling of numeric data and string headers
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            
            # Write existing sheets
            cagr_df.to_excel(writer, sheet_name="Peer Growth Summary")
            comparison_df.to_excel(writer, sheet_name="Comparison Peer Relevant Data")
            
            # Write NEW sheets
            if not price_history_df.empty:
                price_history_df.to_excel(writer, sheet_name="Price History")
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name="Price Summary")

            # Write Peer Summary Statistics sheet (always last)
            if not peer_summary_df.empty:
                sheet_name_benchmark = "Peer Summary Statistics"
                peer_summary_df.to_excel(writer, sheet_name=sheet_name_benchmark, index=False)

                # Apply bold formatting to group title rows in the peer summary sheet
                wb = writer.book
                ws = wb[sheet_name_benchmark]
                group_titles = [g.get("group") for g in PEER_SUMMARY_CONFIG if g.get("group")]
                bold_font = Font(bold=True)
                for row_idx in range(2, ws.max_row + 1):  # skip header row
                    left_cell = ws.cell(row=row_idx, column=1)
                    right_cell = ws.cell(row=row_idx, column=6)
                    if left_cell.value in group_titles:
                        left_cell.font = bold_font
                    if right_cell.value in group_titles:
                        right_cell.font = bold_font
                # Build benchmark lookup dictionaries (relevant and 3Y CAGR means) from Peer Summary Statistics
                benchmark_relevant = {}
                benchmark_cagr = {}
                try:
                    for group_cfg in PEER_SUMMARY_CONFIG:
                        metrics_cfg = group_cfg.get("metrics") or []
                        for metric_cfg in metrics_cfg:
                            metric_name = metric_cfg.get("name")
                            if not metric_name:
                                continue
                            mask = peer_summary_df["Metric"] == metric_name
                            if mask.any():
                                row = peer_summary_df.loc[mask].iloc[0]
                                benchmark_relevant[metric_name] = row.get("Relevant Mean of ALL Stocks", np.nan)
                                benchmark_cagr[metric_name] = row.get("3Y CAGR Mean of ALL Stocks", np.nan)
                except Exception:
                    # If anything goes wrong building the benchmark maps, fall back to empty dicts
                    benchmark_relevant = {}
                    benchmark_cagr = {}

                # Build the "Benchmark Comparison" sheet (per‑ticker % diff and CAGR % diff vs benchmark)
                                # Build the "Benchmark Comparison" sheet (per-ticker % diff and CAGR % diff vs benchmark)
                try:
                    benchmark_sheet_name = "Benchmark Comparison"
                    wb_bench = writer.book
                    if benchmark_sheet_name in wb_bench.sheetnames:
                        ws_bench = wb_bench[benchmark_sheet_name]
                    else:
                        ws_bench = wb_bench.create_sheet(title=benchmark_sheet_name)

                    # Prepare layout rows so EACH ticker section repeats the full metric list
                    layout_rows = []
                    for group_cfg in PEER_SUMMARY_CONFIG:
                        group_name = group_cfg.get("group", "")
                        metrics_cfg = group_cfg.get("metrics") or []

                        # Spacer-only group
                        if not group_name and not metrics_cfg:
                            layout_rows.append({"label": "", "metric_name": None, "use_cagr": False})
                            continue

                        # Group title row
                        if group_name:
                            layout_rows.append({"label": group_name, "metric_name": None, "use_cagr": False})

                        # Metric rows within this group
                        for metric_cfg in metrics_cfg:
                            m_name = metric_cfg.get("name")
                            if not m_name:
                                continue
                            use_cagr = bool(metric_cfg.get("cagr", True))
                            layout_rows.append({
                                "label": m_name,
                                "metric_name": m_name,
                                "use_cagr": use_cagr,
                            })

                    base_row = 2  # first data row (row 1 = headers)
                    metrics_section_rows = len(layout_rows)
                    cagr_cache = {}

                    # For each ticker, create a block: Metric | %Diff | CAGR %Diff | spacer
                    for block_idx, ticker in enumerate(TICKERS):
                        col_start = 1 + block_idx * 6  # 3 used cols + 1 spacer between tickers

                        # Headers for this ticker block
                        ws_bench.cell(row=1, column=col_start).value = "Metric"
                        ws_bench.cell(row=1, column=col_start + 1).value = f"{ticker} Relevant Data"
                        ws_bench.cell(row=1, column=col_start + 2).value = f"{ticker} CAGR (Most Recent / 3Y)"
                        ws_bench.cell(row=1, column=col_start + 3).value = f"{ticker} % Difference from AVR"
                        ws_bench.cell(row=1, column=col_start + 4).value = f"{ticker} CAGR % Difference"

                        # Fill the metric rows
                        for offset, row_info in enumerate(layout_rows):
                            row_idx2 = base_row + offset
                            label = row_info["label"]
                            metric_name = row_info["metric_name"]
                            use_cagr = row_info["use_cagr"]

                            # Metric or group label
                            ws_bench.cell(row=row_idx2, column=col_start).value = label

                            if metric_name:
                                # --- Pull ticker's relevant value from Comparison Peer Relevant Data ---
                                metric_val = np.nan
                                if (metric_name in comparison_df.index) and (ticker in comparison_df.columns):
                                    try:
                                        metric_val = float(comparison_df.at[metric_name, ticker])
                                    except Exception:
                                        metric_val = np.nan
                            
                                # Write raw relevant data value into the Benchmark Comparison sheet
                                if isinstance(metric_val, (int, float)) and not np.isnan(metric_val):
                                    ws_bench.cell(row=row_idx2, column=col_start + 1).value = metric_val
                            
                                # --- Relevant % difference vs benchmark mean ---
                                bench_val = benchmark_relevant.get(metric_name, np.nan)
                                if (isinstance(metric_val, (int, float)) and isinstance(bench_val, (int, float))
                                        and not np.isnan(metric_val) and not np.isnan(bench_val) and bench_val != 0):
                                    abs_diff = abs(metric_val - bench_val) / abs(bench_val) * 100.0
                                    if metric_val > bench_val:
                                        diff_val = abs_diff
                                    elif metric_val < bench_val:
                                        diff_val = -abs_diff
                                    else:
                                        diff_val = 0.0
                                    ws_bench.cell(row=row_idx2, column=col_start + 3).value = diff_val
                            
                                # --- CAGR value and % difference vs benchmark CAGR mean ---
                                if use_cagr:
                                    # Benchmark CAGR mean from Peer Summary Statistics (unchanged)
                                    bench_cagr_val = benchmark_cagr.get(metric_name, np.nan)

                                    # Ticker's CAGR (Most Recent / 3Y) pulled from Peer Growth Summary
                                    ticker_cagr_val = _get_peer_growth_cagr(metric_name, ticker)

                                    # Write raw ticker CAGR (Most Recent / 3Y) value
                                    if isinstance(ticker_cagr_val, (int, float)) and not np.isnan(ticker_cagr_val):
                                        ws_bench.cell(row=row_idx2, column=col_start + 2).value = ticker_cagr_val

                                    # CAGR % difference vs benchmark CAGR mean (signed, absolute base)
                                    if (
                                        isinstance(ticker_cagr_val, (int, float))
                                        and isinstance(bench_cagr_val, (int, float))
                                        and not np.isnan(ticker_cagr_val)
                                        and not np.isnan(bench_cagr_val)
                                        and bench_cagr_val != 0
                                    ):
                                        cagr_diff_val = (ticker_cagr_val - bench_cagr_val) / abs(bench_cagr_val) * 100.0
                                        ws_bench.cell(row=row_idx2, column=col_start + 4).value = cagr_diff_val
                        # Tally rows a couple of rows below the last metric row (per ticker)
                        tally_start = base_row + metrics_section_rows + 2
                        row_gt = tally_start
                        row_lt = tally_start + 1

                        gt_relevant = lt_relevant = 0
                        gt_cagr = lt_cagr = 0

                        for r in range(base_row, base_row + metrics_section_rows):
                            # Relevant column
                            val_rel = ws_bench.cell(row=r, column=col_start + 3).value
                            if isinstance(val_rel, (int, float)):
                                if val_rel > 0:
                                    gt_relevant += 1
                                elif val_rel < 0:
                                    lt_relevant += 1

                            # CAGR column
                            val_cagr = ws_bench.cell(row=r, column=col_start + 4).value
                            if isinstance(val_cagr, (int, float)):
                                if val_cagr > 0:
                                    gt_cagr += 1
                                elif val_cagr < 0:
                                    lt_cagr += 1

                        # Write the tally labels and values under EACH ticker's two data columns
                        ws_bench.cell(row=row_gt, column=col_start).value = "# of > 0"
                        ws_bench.cell(row=row_gt, column=col_start + 3).value = gt_relevant
                        ws_bench.cell(row=row_gt, column=col_start + 4).value = gt_cagr

                        ws_bench.cell(row=row_lt, column=col_start).value = "# of < 0"
                        ws_bench.cell(row=row_lt, column=col_start + 3).value = lt_relevant
                        ws_bench.cell(row=row_lt, column=col_start + 4).value = lt_cagr

                except Exception as e:
                    print(f"Warning: Could not build Benchmark Comparison sheet: {e}")

                # === Fair Value Screen Sheet ===
                try:
                    # ---- Weighted Valuation Scoring ----
                    valuation_metrics = {
                        "EV / EBITDA": {
                            "row_name": "EV / EBITDA",
                            "direction": "lower_better",
                            "weight": 3
                            },
                        "PEG Ratio (Calculated)":{
                            "row_name": "PEG Ratio (Calculated)",
                            "direction": "lower_better",
                            "weight": 3
                            },
                        "EV / FCF Forward (3Y)":{
                            "row_name": "EV / FCF Forward (3Y)",
                            "direction": "lower_better",
                            "weight": 4
                         },
                        "Gross Profit / EV":{
                            "row_name": "Gross Profit / EV",
                            "direction": "higher_better",
                            "weight": 3
                            },
                        "Price / Sales":{
                            "row_name": "Price / Sales",
                            "direction": "lower_better",
                            "weight": 2
                            },
                        "Value Spread (Heavy) (%)": {
                            "row_name": "Value Spread (Heavy) (%)",
                            "direction": "higher_better",
                            "weight": 2
                            },
                        "Value Spread (Light) (%)": {
                            "row_name": "Value Spread (Light) (%)",
                            "direction": "higher_better",
                            "weight": 2
                        },
                    }
                
                    # ---- Weighted Valuation Scoring (Benchmark Relative) ----
                    valuation_rows = []
                    
                    for ticker in TICKERS:
                        if ticker not in comparison_df.columns:
                            continue
                    
                        row_data = {"Ticker": ticker}
                        score_sum = 0
                        weight_sum = 0
                    
                        for metric_name, cfg in valuation_metrics.items():
                            weight = cfg["weight"]
                            direction = cfg["direction"]
                    
                            ticker_val = np.nan
                            if metric_name in comparison_df.index:
                                try:
                                    ticker_val = float(comparison_df.at[metric_name, ticker])
                                except:
                                    ticker_val = np.nan
                    
                            bench_val = benchmark_relevant.get(metric_name, np.nan)
                    
                            if (
                                isinstance(ticker_val, (int, float)) and
                                isinstance(bench_val, (int, float)) and
                                not np.isnan(ticker_val) and
                                not np.isnan(bench_val)
                            ):
                                if direction == "lower_better":
                                    score = -((ticker_val - bench_val) * weight)
                                elif direction == "higher_better":
                                    score = ((ticker_val - bench_val) * weight)
                    
                                row_data[f"{metric_name} Score"] = score
                                score_sum += score
                                weight_sum += weight
                            else:
                                row_data[f"{metric_name} Score"] = np.nan

                        row_data["Fair Value Score"] = score_sum / weight_sum if weight_sum > 0 else np.nan
                        valuation_rows.append(row_data)

                    if valuation_rows:
                        valuation_df = pd.DataFrame(valuation_rows).set_index("Ticker")
                        # Normalize Fair Value Score to 0–100 scale
                        fv = valuation_df["Fair Value Score"]
                        min_val = fv.min()
                        max_val = fv.max()
                        
                        valuation_df["Fair Value Score (0-100)"] = (
                            (fv - min_val) / (max_val - min_val) * 100
                        ).round(2)
                        
                        # Rank based on normalized scores
                        valuation_df["Rank"] = (
                            valuation_df["Fair Value Score (0-100)"]
                            .rank(ascending=False, method="min")
                            .astype("Int64")
                        )
                        # Output cleaned result
                        valuation_df = clean_dataframe_for_export(valuation_df)
                        valuation_df.to_excel(writer, sheet_name="Fair Value Screen", index=True)

                    
                        
                except Exception as e:
                    print(f"Warning: Could not build Fair Value Screen sheet: {e}")


        print(f"Peer Comparison Book saved: {file_path}")


        # NEW: Append WACC assumptions to the Comparison sheet
        try:
            if os.path.exists(file_path):
                book = load_workbook(file_path)
                if "Comparison Peer Relevant Data" in book.sheetnames:
                    ws = book["Comparison Peer Relevant Data"]
                    max_row = ws.max_row + 2
                    ws[f"A{max_row}"] = "MARKET ASSUMPTIONS FOR WACC"
                    max_row += 1
                    ws[f"A{max_row}"] = "Risk-Free Rate"
                    ws[f"B{max_row}"] = RISK_FREE_RATE
                    max_row += 1
                    ws[f"A{max_row}"] = "Equity Risk Premium"
                    ws[f"B{max_row}"] = EQUITY_RISK_PREMIUM
                    max_row += 1
                    ws[f"A{max_row}"] = "Minimum Cost of Debt"
                    ws[f"B{max_row}"] = MIN_COST_OF_DEBT
                    book.save(file_path)
        except Exception as e:
            print(f"Warning: Could not write WACC assumptions to comparison book: {e}")
        
        # Automatic opening of the file
        if os.path.exists(file_path):
            print(f"Attempting to automatically open the Excel file: {file_path}")
            try:
                os.system(f'open "{file_path}"')
            except Exception as e:
                print(f"Warning: Could not automatically open {file_path}. Error: {e}")

    except Exception as e:
        print(f"\nCRITICAL CONSOLIDATED COMPARISON SAVE ERROR: An error occurred while writing the Comparison book: {e}")


# =========================================================================
# === MAIN EXECUTION BLOCK (CORRECTED) ===
# =========================================================================

# --- Hardcoded Ticker List (UNCHANGED) ---
TICKERS = [
    "AMD", "NVDA", "INTC", "QCOM", "TXN", "AVGO", "MU", "SMCI", "MRVL", "LRCX",
    "AMAT", "KLAC", "ASML", "TSM", "UMC", "STM", "INFY", "ON", "MCHP", "LSCC",
    "VECO", "TER", "UCTT", "ENOV", "AEIS", "COHU", "FORM", "NVMI", "WOLF", "POWI",
    "SITM"
]

if __name__ == "__main__":
    
    HISTORY_PATH = "data/subindustry_semiconductors.csv"
    TODAY = pd.Timestamp.today().strftime("%Y-%m-%d")
    
    if not os.path.exists(HISTORY_PATH):
        history_df = bootstrap_subindustry_history()
    else:
        history_df = pd.read_csv(HISTORY_PATH)
    
    missing_dates = detect_missing_trading_days(history_df, TODAY)
    
    if missing_dates:
        history_df = backfill_missing_days(
            missing_dates,
            compute_subindustry_snapshot,
            history_df
        )

    history_df.to_csv(HISTORY_PATH, index=False)

    # NEW: Get input for the base file name
    name_input = input("Enter the base name for the Excel files (e.g., Semiconductor): ").strip()
    if not name_input:
        name_input = "Financials" # Default fallback
        print(f"No name entered. Using default name: {name_input}")
    
    start_time = time.time()
    
    # Containers to hold results for consolidation
    consolidated_raw_dfs = {}
    consolidated_calcs = {}
    static_results_map = {}
    shares_outstanding_map = {} # New map for price metrics calculation

    # --- 0. Initial Setup ---
    print("=======================================================")
    print(f"SEC Financials Scraper Initialized. Tickers to process: {len(TICKERS)}")
    print("=======================================================")

    # --- 1. Loop Through All Tickers to Fetch Data ---
    for ticker in TICKERS:
        print(f"\n\n=======================================================")
        print(f"--- STARTING DATA FETCH FOR {ticker} ---")
        print(f"=======================================================")
        
        # 1. Fetch and Calculate Data
        raw_df, calculated_metrics_df, static_results = fetch_financial_data(ticker)
        
        if raw_df is not None and calculated_metrics_df is not None and not calculated_metrics_df.empty:
            
            # 2. Store DataFrames for Consolidation
            consolidated_raw_dfs[ticker] = raw_df
            consolidated_calcs[ticker] = calculated_metrics_df
            static_results_map[ticker] = static_results
            
            # NEW: Extract most recent Shares Outstanding (Fixed logic from previous step)
            most_recent_shares = 0.0
            if 'Total Common Shares Outstanding' in calculated_metrics_df.index:
                # Find the first (most recent) non-NaN value
                shares_row = calculated_metrics_df.loc['Total Common Shares Outstanding',:]
                try:
                    # Look up the first non-NaN value in the row (which is the most recent year)
                    most_recent_shares = shares_row.dropna().iloc[0]
                except IndexError:
                    pass
            shares_outstanding_map[ticker] = most_recent_shares
            
        else:
            print(f"Skipping consolidation for {ticker}. No financial data found.")
            shares_outstanding_map[ticker] = 0.0

    # Rebuild shares_outstanding_map using RAW data to ensure we have
    # the latest non-NaN Total Common Shares Outstanding for each ticker.
    # This avoids issues where the calculated sheet may contain NaNs for this row.
    for ticker, raw_df in consolidated_raw_dfs.items():
        if raw_df is None or raw_df.empty:
            continue
        try:
            raw_df.columns = raw_df.columns.astype(str)
            if "Total Common Shares Outstanding" in raw_df.index:
                shares_row = raw_df.loc["Total Common Shares Outstanding", :]
                non_na_shares = shares_row.dropna()
                if not non_na_shares.empty:
                    shares_outstanding_map[ticker] = float(non_na_shares.iloc[0])
        except Exception as e:
            print(f"Warning: could not refresh raw shares for {ticker}: {e}")

    # 4. New step: Calculate Price Metrics and Beta
    price_history_df, summary_df = calculate_price_metrics_and_beta(TICKERS, shares_outstanding_map)

    print("\n=======================================================")
    print("--- ALL TICKERS PROCESSED. CREATING CONSOLIDATED BOOKS ---")
    
    # --- STEP A: Create the ALL FINANCIALS Book (New Requirement) ---
    # CORRECTED LINE: Using f-string variable from user input
    all_financials_file_path = f"{name_input}_Full_Data.xlsx"
    create_and_save_all_financials_book(
        consolidated_raw_dfs,
        consolidated_calcs,
        static_results_map,
        all_financials_file_path
    )

    # --- STEP B: Create the PEER COMPARISON Book (Original Requirement) ---
    # CORRECTED LINE: Using f-string variable from user input
    consolidated_file_path = f"{name_input}_Comparison_Data.xlsx"
    
    # Check that we have content for both the old and new sheets before saving
    if consolidated_calcs and not price_history_df.empty and not summary_df.empty:
        # Pass ALL dataframes to the save function
        create_and_save_comparison_book(
            consolidated_calcs,
            price_history_df,
            summary_df,
            consolidated_file_path
        )
    else:
        print("Not enough data to create the Consolidated Peer Comparison book.")
        
    end_time = time.time()
    total_time = end_time - start_time
    print(f"\n--- Script finished in {total_time:.2f} seconds ---")















